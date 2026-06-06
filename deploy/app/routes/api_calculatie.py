"""
Flask Blueprint: Calculatie API
Alle endpoints voor het Maven Calculatiemodel.
"""

import csv
import io
import json
from dataclasses import asdict
from flask import Blueprint, request, jsonify, Response, abort

from ..calculatie import (
    TitelInput, StaffelTrede, KostenPost, DrukConfig, ExtraDerde,
    bereken_titel,
    KanaalResultaat, DrukResultaat, CalculatieResultaat,
)
from .. import storage_calculatie as storage

bp = Blueprint("api_calculatie", __name__)


# ──────────────────────────────────────────────────────────────────
#  BRIDGE: dict → dataclass (vervangt Pydantic + bridge.py)
# ──────────────────────────────────────────────────────────────────

def _staffel_list(items: list[dict]) -> list[StaffelTrede]:
    return [StaffelTrede(s["tot_exemplaren"], s["percentage"]) for s in items]


def _kostenposten_list(items: list[dict]) -> list[KostenPost]:
    return [
        KostenPost(
            id=kp["id"], naam=kp["naam"],
            categorie=kp["categorie"],
            bedrag=kp.get("bedrag", 0.0),
        )
        for kp in items
    ]


def _extra_derden_list(items: list[dict]) -> list[ExtraDerde]:
    return [
        ExtraDerde(
            id=ed.get("id", ""),
            naam=ed.get("naam", ""),
            type=ed.get("type", "royalty"),
            percentage=ed.get("percentage", 0.0),
            staffel=_staffel_list(ed.get("staffel", [])),
            # Voorschot is alleen relevant bij royalty; bij winstdeling
            # negeren we het (al zou de UI 't ook moeten verbergen).
            voorschot=ed.get("voorschot", 0.0) if ed.get("type", "royalty") == "royalty" else 0.0,
        )
        for ed in items
    ]


def _drukken_list(items: list[dict]) -> list[DrukConfig]:
    result = []
    for i, d in enumerate(items):
        result.append(DrukConfig(
            druknummer=d.get("druknummer", i + 1),
            oplage=d.get("oplage", 2000),
            drukkosten_per_ex=d.get("drukkosten_per_ex", 1.20),
            kostenposten=_kostenposten_list(d.get("kostenposten", [])),
            cac_per_ex=d.get("cac_per_ex", 0.0),
        ))
    return result


def dict_to_titel_input(d: dict) -> TitelInput:
    """Converteer JSON dict → TitelInput dataclass."""
    return TitelInput(
        titel=d.get("titel", "Nieuwe titel"),
        auteur=d.get("auteur", ""),
        isbn=d.get("isbn", ""),
        verschijningsdatum=d.get("verschijningsdatum", ""),
        verschenen=d.get("verschenen", False),
        verkoopprijs_incl_btw=d.get("verkoopprijs_incl_btw", 20.0),
        btw_percentage=d.get("btw_percentage", 0.09),
        boekhandelskorting=d.get("boekhandelskorting", 0.48),
        drukken=_drukken_list(d.get("drukken", [])),
        # Webshop
        transactiekosten_pct=d.get("transactiekosten_pct", 0.002),
        fulfillment_per_ex=d.get("fulfillment_per_ex", 4.50),
        cac_per_ex=d.get("cac_per_ex", 0.0),
        # Retail
        distributie_cb_per_ex=d.get("distributie_cb_per_ex", 1.10),
        # B2B
        b2b_porto_per_ex=d.get("b2b_porto_per_ex", 0.0),
        b2b_korting_pct=d.get("b2b_korting_pct", 0.0),
        # Auteur
        auteur_winstdeling_pct=d.get("auteur_winstdeling_pct", 0.0),
        auteur_royalty_staffel=_staffel_list(d.get("auteur_royalty_staffel", [])),
        auteur_voorschot=d.get("auteur_voorschot", 0.0),
        # Derden
        agent_staffel=_staffel_list(d.get("agent_staffel", [])),
        agent_pct=d.get("agent_pct", 0.0),
        agent_winstdeling_pct=d.get("agent_winstdeling_pct", 0.0),
        agent_voorschot=d.get("agent_voorschot", 0.0),
        vertaler_pct=d.get("vertaler_pct", 0.0),
        vertaler_staffel=_staffel_list(d.get("vertaler_staffel", [])),
        vertaler_winstdeling_pct=d.get("vertaler_winstdeling_pct", 0.0),
        vertaler_voorschot=d.get("vertaler_voorschot", 0.0),
        illustrator_pct=d.get("illustrator_pct", 0.0),
        illustrator_staffel=_staffel_list(d.get("illustrator_staffel", [])),
        illustrator_winstdeling_pct=d.get("illustrator_winstdeling_pct", 0.0),
        illustrator_voorschot=d.get("illustrator_voorschot", 0.0),
        # Extra derden (flexibel)
        extra_derden=_extra_derden_list(d.get("extra_derden", [])),
        # Partnership
        heeft_partner=d.get("heeft_partner", False),
        partner_naam=d.get("partner_naam", ""),
        partner_winstdeling_pct=d.get("partner_winstdeling_pct", 0.5),
        # Overig
        overige_kosten_pct=d.get("overige_kosten_pct", 0.0),
    )


def kanaal_to_dict(k: KanaalResultaat) -> dict:
    return asdict(k)


def druk_to_dict(d: DrukResultaat, verd_ws: float, verd_rt: float, verd_b2b: float) -> dict:
    """Converteer DrukResultaat → dict met gewogen marge."""
    ws = kanaal_to_dict(d.webshop)
    rt = kanaal_to_dict(d.retail)
    b2b = kanaal_to_dict(d.b2b)

    gewogen_winst = (
        ws["netto_winst_maven"] * verd_ws
        + rt["netto_winst_maven"] * verd_rt
        + b2b["netto_winst_maven"] * verd_b2b
    )
    gewogen_omzet = (
        ws["netto_omzet"] * verd_ws
        + rt["netto_omzet"] * verd_rt
        + b2b["netto_omzet"] * verd_b2b
    )
    gewogen_marge = gewogen_winst / gewogen_omzet if gewogen_omzet > 0 else 0

    return {
        "druk_type": d.druk_type,
        "oplage": d.oplage,
        "cumulatief_voor_druk": d.cumulatief_voor_druk,
        "kosten_totaal": d.kosten_totaal,
        "webshop": ws,
        "retail": rt,
        "b2b": b2b,
        "gewogen_netto_winst": gewogen_winst,
        "gewogen_netto_omzet": gewogen_omzet,
        "gewogen_marge_pct": gewogen_marge,
    }


def run_calculation(data: dict) -> dict:
    """Voer de volledige calculatie uit vanuit een JSON request dict."""
    ti = data["titel_input"]
    t = dict_to_titel_input(ti)
    verd_ws = data.get("verdeling_webshop", 0.10)
    verd_rt = data.get("verdeling_retail", 0.85)
    verd_b2b = data.get("verdeling_b2b", 0.05)

    res = bereken_titel(t)

    drukken_out = [
        {
            **druk_to_dict(d, verd_ws, verd_rt, verd_b2b),
            "kosten_totaal": d.kosten_totaal,
            "drukkosten_totaal": d.drukkosten_totaal,
        }
        for d in res.drukken
    ]

    # Gewogen marge over ALLE drukken: som euro-winst / som euro-omzet,
    # gewogen met oplage per druk.
    total_winst = sum(d["gewogen_netto_winst"] * d["oplage"] for d in drukken_out)
    total_omzet = sum(d["gewogen_netto_omzet"] * d["oplage"] for d in drukken_out)
    marge_totaal = total_winst / total_omzet if total_omzet > 0 else 0

    return {
        "titel": res.titel,
        "drukken": drukken_out,
        "gewogen_marge_pct_totaal": marge_totaal,
        "totaal_oplage": sum(d["oplage"] for d in drukken_out),
    }


# ──────────────────────────────────────────────────────────────────
#  API ENDPOINTS
# ──────────────────────────────────────────────────────────────────

@bp.route("/api/health")
def health():
    return jsonify(status="ok")


@bp.route("/api/health/storage")
def health_storage():
    """Diagnose-endpoint voor storage-gezondheid.

    Antwoord op: zit er een persistent volume? Hoeveel titels en backups?
    Wanneer was de laatste write? Zo kan een wipe sneller worden opgemerkt.
    """
    import os as _os
    from datetime import datetime as _dt
    db_url = _os.environ.get("DATABASE_URL", "")
    # Maskeer wachtwoord uit DATABASE_URL voor de respons
    db_url_safe = db_url
    if "@" in db_url and "://" in db_url:
        scheme, rest = db_url.split("://", 1)
        if "@" in rest:
            creds, host = rest.split("@", 1)
            db_url_safe = f"{scheme}://***:***@{host}"

    info: dict = {
        "volume_mount_path_env": _os.environ.get("RAILWAY_VOLUME_MOUNT_PATH") or None,
        "data_dir": str(storage.DATA_DIR),
        "titels_file": str(storage.TITELS_FILE),
        "titels_file_exists": storage.TITELS_FILE.exists(),
        "railway_env_vars": {k: v for k, v in _os.environ.items() if k.startswith("RAILWAY_")},
        "data_mount_exists": _os.path.exists("/data"),
        "data_mount_writable": _os.access("/data", _os.W_OK) if _os.path.exists("/data") else False,
        "database_url_set": bool(db_url),
        "database_url_safe": db_url_safe or None,
    }
    try:
        all_data = storage.load_all()
        info["titels_count"] = len(all_data)
        info["titel_names"] = sorted({
            v.get("titel_input", {}).get("titel", "?") for v in all_data.values()
        })
        info["storage_backend"] = "postgres" if db_url.startswith("postgresql") else "sqlite"
    except Exception as exc:
        info["load_error"] = str(exc)
        info["storage_backend"] = "ERROR"

    if storage.TITELS_FILE.exists():
        try:
            stat = storage.TITELS_FILE.stat()
            info["titels_file_size"] = stat.st_size
            info["titels_file_modified"] = _dt.fromtimestamp(stat.st_mtime).isoformat()
        except OSError as exc:
            info["stat_error"] = str(exc)

    try:
        backups = storage.list_backups()
        info["backups_count"] = len(backups)
        if backups:
            info["latest_backup"] = backups[0]
            info["oldest_backup"] = backups[-1]
            info["max_backup_size"] = max(b["size"] for b in backups)
    except Exception as exc:
        info["backup_error"] = str(exc)

    # Eventuele .corrupt-bestanden tonen (geeft signaal dat er ooit
    # iets fout ging en data via die fallback is gered)
    try:
        corrupt = sorted(storage.DATA_DIR.glob("*.corrupt-*"))
        info["corrupt_files"] = [c.name for c in corrupt]
    except Exception:
        info["corrupt_files"] = []

    return jsonify(info)


@bp.route("/api/calculate", methods=["POST"])
def calculate():
    data = request.get_json()
    result = run_calculation(data)
    return jsonify(result)


@bp.route("/api/sensitivity/cac", methods=["POST"])
def sensitivity_cac():
    data = request.get_json()
    cac_range = data.get("cac_range", [0, 1, 2, 3, 4, 5, 6, 7, 8, 10, 12, 15])
    results = []

    for cac_val in cac_range:
        # CAC zit nu per druk; zet voor de sensitivity de CAC op ALLE drukken
        # zodat elke druk in de respons consistent meebeweegt.
        data["titel_input"]["cac_per_ex"] = cac_val
        for druk in data["titel_input"].get("drukken", []) or []:
            druk["cac_per_ex"] = cac_val
        calc = run_calculation(data)

        for druk in calc["drukken"]:
            # Find or create entry for this druk_type
            entry = None
            for r in results:
                if r["druk_type"] == druk["druk_type"]:
                    entry = r
                    break
            if entry is None:
                entry = {
                    "variable_name": "cac_per_ex",
                    "druk_type": druk["druk_type"],
                    "rows": [],
                }
                results.append(entry)

            entry["rows"].append({
                "variable_value": cac_val,
                "webshop_winst": druk["webshop"]["netto_winst_maven"],
                "webshop_marge_pct": druk["webshop"]["marge_pct"],
                "retail_winst": druk["retail"]["netto_winst_maven"],
                "retail_marge_pct": druk["retail"]["marge_pct"],
                "b2b_winst": druk["b2b"]["netto_winst_maven"],
                "b2b_marge_pct": druk["b2b"]["marge_pct"],
                "gewogen_winst": druk["gewogen_netto_winst"],
                "gewogen_marge_pct": druk["gewogen_marge_pct"],
            })

    return jsonify(results)


@bp.route("/api/sensitivity/price", methods=["POST"])
def sensitivity_price():
    data = request.get_json()
    current_price = data["titel_input"].get("verkoopprijs_incl_btw", 20.0)
    # Generate retail-friendly prices in ±€3 range
    retail_prices = []
    for p in [
        # Common Dutch retail price points
        9.99, 10.99, 11.99, 12.50, 12.99, 13.99, 14.99, 15.99,
        16.99, 17.50, 17.99, 18.99, 19.99, 20.99, 21.99, 22.50,
        22.99, 24.99, 25.99, 27.50, 27.99, 29.99, 32.50, 34.99,
    ]:
        if current_price - 3.5 <= p <= current_price + 3.5:
            retail_prices.append(p)
    # Ensure current price is included
    if not any(abs(p - current_price) < 0.01 for p in retail_prices):
        retail_prices.append(current_price)
        retail_prices.sort()
    price_range = data.get("price_range", retail_prices)
    results = []

    for price_val in price_range:
        data["titel_input"]["verkoopprijs_incl_btw"] = price_val
        calc = run_calculation(data)

        for druk in calc["drukken"]:
            entry = None
            for r in results:
                if r["druk_type"] == druk["druk_type"]:
                    entry = r
                    break
            if entry is None:
                entry = {
                    "variable_name": "verkoopprijs_incl_btw",
                    "druk_type": druk["druk_type"],
                    "rows": [],
                }
                results.append(entry)

            entry["rows"].append({
                "variable_value": price_val,
                "webshop_winst": druk["webshop"]["netto_winst_maven"],
                "webshop_marge_pct": druk["webshop"]["marge_pct"],
                "retail_winst": druk["retail"]["netto_winst_maven"],
                "retail_marge_pct": druk["retail"]["marge_pct"],
                "b2b_winst": druk["b2b"]["netto_winst_maven"],
                "b2b_marge_pct": druk["b2b"]["marge_pct"],
                "gewogen_winst": druk["gewogen_netto_winst"],
                "gewogen_marge_pct": druk["gewogen_marge_pct"],
            })

    return jsonify(results)


# ── Titels CRUD ──

@bp.route("/api/titels")
def list_titels():
    include_archived = request.args.get("archived") == "true"
    all_data = storage.load_all()
    items = []
    for tid, tdata in all_data.items():
        archived = tdata.get("archived", False)
        if not include_archived and archived:
            continue
        ti = tdata.get("titel_input", {})
        # Bereken gewogen marge als er genoeg data is
        gewogen_marge = None
        try:
            calc_req = {
                "titel_input": ti,
                "verdeling_webshop": tdata.get("verdeling_webshop", 0.10),
                "verdeling_retail": tdata.get("verdeling_retail", 0.85),
                "verdeling_b2b": tdata.get("verdeling_b2b", 0.05),
            }
            res = run_calculation(calc_req)
            gewogen_marge = res.get("gewogen_marge_pct_totaal")
        except Exception:
            pass
        items.append({
            "id": tid,
            "titel": ti.get("titel", ""),
            "auteur": ti.get("auteur", ""),
            "isbn": ti.get("isbn", ""),
            "drukken_count": len(ti.get("drukken", [])),
            "gewogen_marge_pct": gewogen_marge,
            "archived": archived,
            "titelgroep_id": tdata.get("titelgroep_id"),
        })
    return jsonify(items)


@bp.route("/api/titels/<titel_id>")
def get_titel(titel_id):
    data = storage.get_titel(titel_id)
    if data is None:
        abort(404, description="Titel niet gevonden")
    return jsonify({"id": titel_id, **data})


@bp.route("/api/titels", methods=["POST"])
def save_titel():
    data = request.get_json()
    titel_id = data.get("id") or storage.new_id()

    titel_data = {
        "titel_input": data["titel_input"],
        "herdruk_oplages": data.get("herdruk_oplages", []),
        "verdeling_webshop": data.get("verdeling_webshop", 0.10),
        "verdeling_retail": data.get("verdeling_retail", 0.85),
        "verdeling_b2b": data.get("verdeling_b2b", 0.05),
        "titelgroep_id": data.get("titelgroep_id"),
    }

    storage.save_titel(titel_id, titel_data)
    return jsonify({"id": titel_id, **titel_data})


@bp.route("/api/titels/<titel_id>", methods=["DELETE"])
def delete_titel(titel_id):
    if storage.delete_titel(titel_id):
        return jsonify(ok=True)
    abort(404, description="Titel niet gevonden")


@bp.route("/api/titels/<titel_id>/archive", methods=["PATCH"])
def archive_titel(titel_id):
    data = storage.get_titel(titel_id)
    if data is None:
        abort(404, description="Titel niet gevonden")
    data["archived"] = True
    storage.save_titel(titel_id, data)
    return jsonify(ok=True)


@bp.route("/api/titels/<titel_id>/unarchive", methods=["PATCH"])
def unarchive_titel(titel_id):
    data = storage.get_titel(titel_id)
    if data is None:
        abort(404, description="Titel niet gevonden")
    data["archived"] = False
    storage.save_titel(titel_id, data)
    return jsonify(ok=True)


# ── Titelgroepen CRUD ──

@bp.route("/api/titelgroepen", methods=["GET"])
def list_titelgroepen_route():
    return jsonify(storage.list_titelgroepen())


@bp.route("/api/titelgroepen/<groep_id>", methods=["GET"])
def get_titelgroep_route(groep_id):
    with_titels = request.args.get("with_titels") == "true"
    data = storage.get_titelgroep(groep_id, with_titels=with_titels)
    if data is None:
        abort(404, description="Titelgroep niet gevonden")
    return jsonify(data)


@bp.route("/api/titelgroepen", methods=["POST"])
def create_titelgroep_route():
    data = request.get_json() or {}
    if not data.get("naam"):
        abort(400, description="naam is verplicht")
    return jsonify(storage.save_titelgroep(None, data))


@bp.route("/api/titelgroepen/<groep_id>", methods=["PUT", "PATCH"])
def update_titelgroep_route(groep_id):
    data = request.get_json() or {}
    if storage.get_titelgroep(groep_id) is None:
        abort(404, description="Titelgroep niet gevonden")
    return jsonify(storage.save_titelgroep(groep_id, data))


@bp.route("/api/titelgroepen/<groep_id>", methods=["DELETE"])
def delete_titelgroep_route(groep_id):
    if storage.delete_titelgroep(groep_id):
        return jsonify(ok=True)
    abort(404, description="Titelgroep niet gevonden")


@bp.route("/api/backup/export", methods=["GET"])
def backup_export():
    """Download alle titels als JSON-bestand voor handmatige backup."""
    from io import BytesIO
    from datetime import datetime as _dt
    from flask import send_file
    all_data = storage.load_all()
    buf = BytesIO(json.dumps(all_data, ensure_ascii=False, indent=2).encode("utf-8"))
    buf.seek(0)
    ts = _dt.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"calculatie_backup_{ts}.json",
        mimetype="application/json",
    )


@bp.route("/api/backup/list", methods=["GET"])
def backup_list():
    """Lijst beschikbare automatische backups."""
    return jsonify({"backups": storage.list_backups()})


@bp.route("/api/backup/restore", methods=["POST"])
def backup_restore():
    """Restore een eerdere backup over de huidige data heen.

    Body: {"name": "calculatie_titels_20260530_120000.json"}
    De huidige data wordt eerst nog als backup bewaard.
    """
    data = request.get_json() or {}
    name = data.get("name", "")
    if not name:
        abort(400, description="Geen backup-naam meegegeven")
    if storage.restore_backup(name):
        return jsonify({"restored": True, "name": name})
    abort(404, description="Backup niet gevonden")


@bp.route("/api/backup/import", methods=["POST"])
def backup_import():
    """Upload een JSON-backup en merge in de database.

    Body: {"data": {...}, "mode": "merge"|"replace"}
    - merge (default): voeg titels toe, overschrijf bestaande met dezelfde id
    - replace: vervang de hele database (huidige data wordt eerst gebackupt)
    """
    body = request.get_json() or {}
    new_data = body.get("data")
    mode = body.get("mode", "merge")
    if not isinstance(new_data, dict):
        abort(400, description="data moet een object zijn met titel-id's als sleutels")

    count = storage.import_data(new_data, mode=mode)
    if mode == "replace":
        return jsonify({"imported": count, "mode": "replace"})
    total = len(storage.load_all())
    return jsonify({"imported": count, "total": total, "mode": "merge"})


@bp.route("/api/seed", methods=["POST"])
def seed_database():
    """Seed de database met Maven-titels als deze leeg is."""
    from ..seed_data import MAVEN_TITELS
    all_data = storage.load_all()
    if len(all_data) > 0:
        return jsonify({"seeded": 0, "message": "Database is niet leeg"})

    count = 0
    for t in MAVEN_TITELS:
        tid = storage.new_id()
        titel_data = {
            "titel_input": {
                "titel": t["titel"],
                "auteur": t["auteur"],
                "isbn": t["isbn"],
                "verschijningsdatum": "",
                "verschenen": t["druknummer"] > 0,
                "verkoopprijs_incl_btw": 20.0,
                "btw_percentage": 0.09,
                "boekhandelskorting": 0.48,
                "drukken": [{
                    "druknummer": t["druknummer"],
                    "oplage": 2000,
                    "drukkosten_per_ex": 1.20,
                    "kostenposten": [],
                }],
                "transactiekosten_pct": 0.002,
                "fulfillment_per_ex": 4.50,
                "cac_per_ex": 0.0,
                "distributie_cb_per_ex": 1.10,
                "b2b_porto_per_ex": 0.0,
                "b2b_korting_pct": 0.0,
                "auteur_winstdeling_pct": 0.50,
                "auteur_royalty_staffel": [],
                "auteur_voorschot": 0,
                "agent_pct": 0.0, "agent_staffel": [], "agent_winstdeling_pct": 0.0, "agent_voorschot": 0,
                "vertaler_pct": 0.0, "vertaler_staffel": [], "vertaler_winstdeling_pct": 0.0, "vertaler_voorschot": 0,
                "illustrator_pct": 0.0, "illustrator_staffel": [], "illustrator_winstdeling_pct": 0.0, "illustrator_voorschot": 0,
                "heeft_partner": False,
                "partner_naam": "",
                "partner_winstdeling_pct": 0.5,
                "overige_kosten_pct": 0.0,
                "overige_kosten_items": [],
                "extra_derden": [],
            },
            "verdeling_webshop": 0.10,
            "verdeling_retail": 0.85,
            "verdeling_b2b": 0.05,
            "archived": False,
        }
        storage.save_titel(tid, titel_data)
        count += 1

    return jsonify({"seeded": count})


# ── Oplage Simulatie ──

@bp.route("/api/simulate/oplage", methods=["POST"])
def simulate_oplage():
    """Simuleer P&L bij verschillende verkoopaantallen.

    Berekent netto resultaat incl. kostenposten van 1e druk en voorschotten.
    Geeft 4 punten: huidige oplage, break-even, +5000, +10000.
    """
    data = request.get_json()
    ti = data.get("titel_input", {})
    verd_ws = data.get("verdeling_webshop", 0.10)
    verd_rt = data.get("verdeling_retail", 0.85)
    verd_b2b = data.get("verdeling_b2b", 0.05)

    try:
        calc = run_calculation(data)
    except Exception:
        return jsonify({"rows": [], "break_even_oplage": None})

    if not calc["drukken"]:
        return jsonify({"rows": [], "break_even_oplage": None})

    druk = calc["drukken"][0]
    ws, rt, b2b_k = druk["webshop"], druk["retail"], druk["b2b"]

    # Gewogen per-exemplaar values
    netto_omzet_per_ex = ws["netto_omzet"] * verd_ws + rt["netto_omzet"] * verd_rt + b2b_k["netto_omzet"] * verd_b2b
    netto_winst_per_ex = ws["netto_winst_maven"] * verd_ws + rt["netto_winst_maven"] * verd_rt + b2b_k["netto_winst_maven"] * verd_b2b

    # The engine amortizes kostenposten into the per-ex result (kosten_per_ex).
    # For the oplage sim, strip that out and add them back as a fixed lump sum,
    # so the sim can model volumes different from the druk oplage.
    kosten_per_ex_gewogen = (
        ws["kosten_per_ex"] * verd_ws + rt["kosten_per_ex"] * verd_rt + b2b_k["kosten_per_ex"] * verd_b2b
    )
    var_winst_per_ex = netto_winst_per_ex + kosten_per_ex_gewogen

    # Fixed costs: kostenposten van de 1e druk
    totaal_eenmalig = druk.get("kosten_totaal", 0)

    drukken_config = ti.get("drukken", [])
    if not drukken_config:
        drukken_config = [{"druknummer": 1, "oplage": 2000, "drukkosten_per_ex": 1.20}]

    # Total drukkosten as fixed investment
    totaal_drukkosten = sum(d.get("oplage", 0) * d.get("drukkosten_per_ex", 0) for d in drukken_config)
    totaal_oplage = sum(d.get("oplage", 0) for d in drukken_config)

    # Voorschotten per partij. Alleen relevant als er ÓÓK een actieve
    # per-ex stroom is om het voorschot tegen weg te schrijven — anders
    # is het een achtergebleven veld zonder deal en mag het niet meetellen.
    def _heeft_royalty_deal(prefix: str) -> bool:
        return ti.get(f"{prefix}_pct", 0) > 0 or len(ti.get(f"{prefix}_staffel", []) or []) > 0

    auteur_voorschot = ti.get("auteur_voorschot", 0) if len(ti.get("auteur_royalty_staffel") or []) > 0 else 0
    agent_voorschot = ti.get("agent_voorschot", 0) if _heeft_royalty_deal("agent") else 0
    vertaler_voorschot = ti.get("vertaler_voorschot", 0) if _heeft_royalty_deal("vertaler") else 0
    illustrator_voorschot = ti.get("illustrator_voorschot", 0) if _heeft_royalty_deal("illustrator") else 0
    extra_derden_voorschot = sum(
        d.get("voorschot", 0) for d in ti.get("extra_derden", [])
        if d.get("type") == "royalty" and (
            d.get("percentage", 0) > 0 or len(d.get("staffel") or []) > 0
        )
    )
    totaal_voorschotten = (
        auteur_voorschot + agent_voorschot + vertaler_voorschot
        + illustrator_voorschot + extra_derden_voorschot
    )

    # Per-ex royalty/commission per partij (already deducted in netto_winst_maven)
    drukkosten_in_perex = ws["drukkosten"] * verd_ws + rt["drukkosten"] * verd_rt + b2b_k["drukkosten"] * verd_b2b
    pure_var_winst_per_ex = var_winst_per_ex + drukkosten_in_perex  # add back drukkosten

    royalty_per_ex = ws["auteur_royalty"] * verd_ws + rt["auteur_royalty"] * verd_rt + b2b_k["auteur_royalty"] * verd_b2b
    agent_per_ex = ws["agent"] * verd_ws + rt["agent"] * verd_rt + b2b_k["agent"] * verd_b2b
    vertaler_per_ex = ws["vertaler"] * verd_ws + rt["vertaler"] * verd_rt + b2b_k["vertaler"] * verd_b2b
    illustrator_per_ex = ws["illustrator"] * verd_ws + rt["illustrator"] * verd_rt + b2b_k["illustrator"] * verd_b2b

    # Correct per-ex margin: add back the per-ex royalty/commission for each party
    # that has a voorschot (those will be deducted as effective fixed costs instead).
    # Parties WITHOUT a voorschot keep their per-ex cost as an ongoing variable cost.
    adjusted_per_ex = pure_var_winst_per_ex
    if auteur_voorschot > 0:
        adjusted_per_ex += royalty_per_ex
    if agent_voorschot > 0 and agent_per_ex > 0:
        adjusted_per_ex += agent_per_ex
    if vertaler_voorschot > 0 and vertaler_per_ex > 0:
        adjusted_per_ex += vertaler_per_ex
    if illustrator_voorschot > 0 and illustrator_per_ex > 0:
        adjusted_per_ex += illustrator_per_ex

    def calc_result_at_volume(vol):
        """Calculate net result at a given total volume sold."""
        # Drukkosten: sum up per-druk, capped at the volume
        druk_costs = 0
        remaining = vol
        for d in sorted(drukken_config, key=lambda x: x.get("druknummer", 1)):
            druk_vol = min(remaining, d.get("oplage", 0))
            druk_costs += druk_vol * d.get("drukkosten_per_ex", 0)
            remaining -= druk_vol
            if remaining <= 0:
                break
        if remaining > 0 and drukken_config:
            last_druk = drukken_config[-1]
            druk_costs += remaining * last_druk.get("drukkosten_per_ex", 1.20)

        # Effective cost per partij met voorschot:
        #   If earned royalties < voorschot: Maven pays exactly the voorschot (advance not yet recouped).
        #   If earned royalties > voorschot: Maven pays ongoing royalties (advance fully recouped).
        # For parties WITHOUT voorschot: per-ex cost already in adjusted_per_ex (variable).
        effective_fixed = totaal_eenmalig

        if auteur_voorschot > 0:
            auteur_earned = vol * royalty_per_ex
            effective_fixed += max(auteur_voorschot, auteur_earned) if royalty_per_ex > 0 else auteur_voorschot
        if agent_voorschot > 0:
            agent_earned = vol * agent_per_ex
            effective_fixed += max(agent_voorschot, agent_earned) if agent_per_ex > 0 else agent_voorschot
        if vertaler_voorschot > 0:
            vertaler_earned = vol * vertaler_per_ex
            effective_fixed += max(vertaler_voorschot, vertaler_earned) if vertaler_per_ex > 0 else vertaler_voorschot
        if illustrator_voorschot > 0:
            illustrator_earned = vol * illustrator_per_ex
            effective_fixed += max(illustrator_voorschot, illustrator_earned) if illustrator_per_ex > 0 else illustrator_voorschot
        # extra_derden voorschotten: treat as pure fixed costs (no per-ex recoupment tracked)
        effective_fixed += extra_derden_voorschot

        net_result = vol * adjusted_per_ex - druk_costs - effective_fixed

        total_omzet = vol * netto_omzet_per_ex
        marge = net_result / total_omzet if total_omzet > 0 else -10

        # voorschot_ingelopen: are all royalties/commissions >= their respective voorschotten?
        totaal_earned = (
            (vol * royalty_per_ex if auteur_voorschot > 0 else 0)
            + (vol * agent_per_ex if agent_voorschot > 0 else 0)
            + (vol * vertaler_per_ex if vertaler_voorschot > 0 else 0)
            + (vol * illustrator_per_ex if illustrator_voorschot > 0 else 0)
        )
        voorschot_ingelopen = totaal_earned >= totaal_voorschotten if totaal_voorschotten > 0 else True

        return {
            "oplage": vol,
            "omzet": round(total_omzet, 2),
            "kosten": round(total_omzet - net_result, 2),
            "netto_resultaat": round(net_result, 2),
            "marge_pct": round(marge, 4),
            "is_break_even": False,
            "is_voorschot_earn_out": False,
            "voorschot_ingelopen": voorschot_ingelopen,
        }

    # P&L break-even (Maven's netto resultaat = 0, inclusief voorschot als investering)
    def find_break_even():
        r_at_1 = calc_result_at_volume(1)
        r_high = calc_result_at_volume(200000)
        if r_at_1["netto_resultaat"] >= 0:
            return None
        if r_high["netto_resultaat"] < 0:
            return None
        low, high = 1, 200000
        for _ in range(50):
            mid = (low + high) // 2
            if calc_result_at_volume(mid)["netto_resultaat"] < 0:
                low = mid
            else:
                high = mid
            if high - low <= 10:
                break
        return max(((high + 24) // 50) * 50, 50)

    # Voorschot earn-out: oplage waarop de royalty's het voorschot dekken
    def find_earn_out():
        total_recoup_per_ex = (
            (royalty_per_ex if auteur_voorschot > 0 else 0)
            + (agent_per_ex if agent_voorschot > 0 else 0)
            + (vertaler_per_ex if vertaler_voorschot > 0 else 0)
            + (illustrator_per_ex if illustrator_voorschot > 0 else 0)
        )
        active_vs = (
            (auteur_voorschot if auteur_voorschot > 0 else 0)
            + (agent_voorschot if agent_voorschot > 0 else 0)
            + (vertaler_voorschot if vertaler_voorschot > 0 else 0)
            + (illustrator_voorschot if illustrator_voorschot > 0 else 0)
        )
        if active_vs <= 0 or total_recoup_per_ex <= 0:
            return None
        eo = int(active_vs / total_recoup_per_ex) + 1
        return ((eo + 49) // 50) * 50

    break_even = find_break_even()
    voorschot_earn_out = find_earn_out()

    # Build simulation points: break-even, voorschot-earn-out, huidige oplage, +5k, +10k
    volumes = set()
    volumes.add(totaal_oplage)
    if break_even is not None and break_even > 0:
        volumes.add(break_even)
    if voorschot_earn_out is not None and voorschot_earn_out > 0:
        volumes.add(voorschot_earn_out)
    volumes.add(totaal_oplage + 5000)
    volumes.add(totaal_oplage + 10000)

    # Beperken tot maximaal 5 punten (anders te druk)
    sorted_vols = sorted(volumes)[:5]

    rows = []
    for vol in sorted_vols:
        row = calc_result_at_volume(vol)
        if break_even is not None and vol == break_even:
            row["is_break_even"] = True
        if voorschot_earn_out is not None and vol == voorschot_earn_out:
            row["is_voorschot_earn_out"] = True
        rows.append(row)

    return jsonify({
        "rows": rows,
        "break_even_oplage": break_even,
        "voorschot_earn_out_oplage": voorschot_earn_out,
    })


# ── Excel Export ──

@bp.route("/api/export/excel", methods=["POST"])
def export_excel():
    """Exporteer calculatie naar een leesbaar Excel-bestand.

    Twee tabbladen:
    - 'Calculatie' — alle invoer (basisgegevens, drukken, kostenposten,
      verkoopkanalen, deals + voorschotten van alle partijen)
    - 'Resultaat' — gewogen marge, marge per kanaal waterfall (royalty's
      boven brutowinst, winstdelingen eronder), oplage-simulatie met
      voorschot-ingelopen kolom

    Bedoeld als rapportage-document (1-op-1 overname uit de app), niet
    als rekenmachine met formules.
    """
    from io import BytesIO
    from datetime import date as _date
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = request.get_json()
    if not data:
        abort(400, description="Geen data meegegeven")

    ti = data.get("titel_input", {})
    verd_ws = data.get("verdeling_webshop", 0.10)
    verd_rt = data.get("verdeling_retail", 0.85)
    verd_b2b = data.get("verdeling_b2b", 0.05)

    try:
        calc = run_calculation(data)
    except Exception as e:
        abort(400, description=f"Berekening mislukt: {e}")

    wb = openpyxl.Workbook()

    # ── Gedeelde stijlen (consistent over beide tabbladen) ──
    GROEN = "FF1B5E20"
    LICHTGROEN = "FFE8F5E9"
    GRIJS_HEADER = "FF37474F"
    LICHT_GRIJS = "FFEEEEEE"
    WIT = "FFFFFFFF"

    def h1(cell, text):
        cell.value = text
        cell.font = Font(bold=True, size=13, color=WIT)
        cell.fill = PatternFill("solid", fgColor=GRIJS_HEADER)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    def h2(cell, text):
        cell.value = text
        cell.font = Font(bold=True, size=10, color=GROEN)
        cell.fill = PatternFill("solid", fgColor=LICHTGROEN)

    def label(cell, text, bold=False):
        cell.value = text
        cell.font = Font(size=10, bold=bold)
        cell.alignment = Alignment(horizontal="left", indent=1)

    def val(cell, v, fmt=None, bold=False, color=None):
        cell.value = v
        cell.font = Font(size=10, bold=bold, color=color or "FF000000")
        if fmt:
            cell.number_format = fmt
        cell.alignment = Alignment(horizontal="right")

    def th(cell, text):
        """Tabel-kolom-header."""
        cell.value = text
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill("solid", fgColor=LICHT_GRIJS)
        cell.alignment = Alignment(horizontal="right")

    def set_widths(ws, *widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ──────────────────────────────────────────────────────────────────
    #  TAB 1 — CALCULATIE (inputs)
    # ──────────────────────────────────────────────────────────────────
    ws_sheet = wb.active
    ws_sheet.title = "Calculatie"

    set_widths(ws_sheet, 36, 16, 16, 16, 16, 16)
    r = 1

    # ── HEADER ──
    ws_sheet.merge_cells(f"A{r}:F{r}")
    h1(ws_sheet[f"A{r}"], f"Calculatie — {ti.get('titel', '?')}")
    ws_sheet.row_dimensions[r].height = 24
    r += 1
    ws_sheet.merge_cells(f"A{r}:F{r}")
    meta = f"{ti.get('auteur','') or '—'}"
    if ti.get('isbn'):
        meta += f"  ·  ISBN {ti['isbn']}"
    meta += f"  ·  Gegenereerd {_date.today().strftime('%d-%m-%Y')}"
    ws_sheet[f"A{r}"].value = meta
    ws_sheet[f"A{r}"].font = Font(size=9, italic=True, color="FF666666")
    r += 2

    # ── BASISGEGEVENS ──
    ws_sheet.merge_cells(f"A{r}:F{r}")
    h2(ws_sheet[f"A{r}"], "BASISGEGEVENS")
    r += 1

    drukken_cfg = ti.get("drukken", [{"druknummer": 1, "oplage": 2000, "drukkosten_per_ex": 1.20}])
    vkp_incl = ti.get("verkoopprijs_incl_btw", 0)
    btw_pct_v = ti.get("btw_percentage", 0.09)
    bh_korting = ti.get("boekhandelskorting", 0.48)
    vkp_ex = vkp_incl / (1 + btw_pct_v) if btw_pct_v > -1 else vkp_incl

    basis = [
        ("Verkoopprijs incl. BTW", vkp_incl, "€ #,##0.00"),
        ("BTW", btw_pct_v, "0.0%"),
        ("Verkoopprijs ex BTW", vkp_ex, "€ #,##0.00"),
        ("Boekhandelskorting", bh_korting, "0.0%"),
    ]
    for row_label, row_val, row_fmt in basis:
        label(ws_sheet[f"A{r}"], row_label)
        val(ws_sheet[f"B{r}"], row_val, row_fmt)
        r += 1
    r += 1

    # ── DRUKKEN + KOSTENPOSTEN PER DRUK ──
    ws_sheet.merge_cells(f"A{r}:F{r}")
    h2(ws_sheet[f"A{r}"], "DRUKKEN")
    r += 1

    for dk in drukken_cfg:
        druknr = dk.get("druknummer", 1)
        oplage = dk.get("oplage", 0)
        drukkosten_ex = dk.get("drukkosten_per_ex", 0)

        # Sub-header per druk
        ws_sheet.cell(r, 1).value = f"{druknr}e druk"
        ws_sheet.cell(r, 1).font = Font(bold=True, size=10)
        ws_sheet.cell(r, 1).fill = PatternFill("solid", fgColor="FFF5F5F5")
        ws_sheet.cell(r, 2).value = f"{oplage:,}".replace(",", ".") + " exemplaren"
        ws_sheet.cell(r, 2).font = Font(size=9, italic=True, color="FF666666")
        r += 1

        label(ws_sheet[f"A{r}"], "  Drukkosten per ex")
        val(ws_sheet[f"B{r}"], drukkosten_ex, "€ #,##0.00")
        r += 1
        label(ws_sheet[f"A{r}"], "  Drukkosten totaal")
        val(ws_sheet[f"B{r}"], oplage * drukkosten_ex, "€ #,##0")
        r += 1

        # Kostenposten gegroepeerd per categorie
        kp_per_cat: dict[str, list] = {}
        for kp in dk.get("kostenposten", []):
            if (kp.get("bedrag") or 0) <= 0:
                continue
            kp_per_cat.setdefault(kp.get("categorie", "overig"), []).append(kp)
        cat_labels = {
            "productie": "Productie",
            "offline_marketing": "Offline marketing",
            "online_marketing": "Online marketing",
        }
        for cat_key, cat_label in cat_labels.items():
            items = kp_per_cat.get(cat_key, [])
            if not items:
                continue
            label(ws_sheet[f"A{r}"], f"  {cat_label}", bold=True)
            r += 1
            for kp in items:
                label(ws_sheet[f"A{r}"], f"    {kp.get('naam', kp.get('id', '?'))}")
                eur_fmt = "€ #,##0"
                val(ws_sheet[f"B{r}"], kp.get("bedrag", 0), eur_fmt)
                r += 1

        # CAC voor deze druk
        cac = dk.get("cac_per_ex", 0)
        if cac:
            label(ws_sheet[f"A{r}"], "  CAC (online ads per webshop-aankoop)")
            val(ws_sheet[f"B{r}"], cac, "€ #,##0.00")
            r += 1
        r += 1

    # ── VERKOOPKANALEN ──
    ws_sheet.merge_cells(f"A{r}:F{r}")
    h2(ws_sheet[f"A{r}"], "VERKOOPKANALEN")
    r += 1
    label(ws_sheet[f"A{r}"], "Verdeling retail / CB")
    val(ws_sheet[f"B{r}"], verd_rt, "0.0%")
    r += 1
    label(ws_sheet[f"A{r}"], "Verdeling webshop")
    val(ws_sheet[f"B{r}"], verd_ws, "0.0%")
    r += 1
    label(ws_sheet[f"A{r}"], "Verdeling B2B")
    val(ws_sheet[f"B{r}"], verd_b2b, "0.0%")
    r += 1

    kanaal_inputs = [
        ("Boekhandelskorting (retail/CB)", bh_korting, "0.0%"),
        ("CB-distributie per ex (retail/CB)", ti.get("distributie_cb_per_ex", 1.10), "€ #,##0.00"),
        ("Fulfillment per ex (webshop)", ti.get("fulfillment_per_ex", 4.50), "€ #,##0.00"),
        ("Transactiekosten (webshop)", ti.get("transactiekosten_pct", 0.002), "0.0%"),
        ("B2B korting", ti.get("b2b_korting_pct", 0), "0.0%"),
        ("B2B porto per ex", ti.get("b2b_porto_per_ex", 0), "€ #,##0.00"),
    ]
    for row_label, row_val, row_fmt in kanaal_inputs:
        label(ws_sheet[f"A{r}"], row_label)
        val(ws_sheet[f"B{r}"], row_val, row_fmt)
        r += 1
    r += 1

    # ── DEALS ──
    ws_sheet.merge_cells(f"A{r}:F{r}")
    h2(ws_sheet[f"A{r}"], "DEALS & VOORSCHOTTEN")
    r += 1

    # Header
    for col, hdr in enumerate(["Partij", "Type", "Percentage", "Voorschot"], 1):
        th(ws_sheet.cell(r, col), hdr)
        ws_sheet.cell(r, col).alignment = Alignment(horizontal="left" if col == 1 else "right")
    r += 1

    def deal_row(partij: str, type_str: str, pct_val, voorschot_val):
        ws_sheet.cell(r, 1).value = partij
        ws_sheet.cell(r, 1).font = Font(size=10, bold=True)
        ws_sheet.cell(r, 2).value = type_str
        ws_sheet.cell(r, 2).font = Font(size=10)
        if pct_val is None:
            ws_sheet.cell(r, 3).value = "—"
            ws_sheet.cell(r, 3).alignment = Alignment(horizontal="right")
        else:
            val(ws_sheet.cell(r, 3), pct_val, "0.0%")
        if voorschot_val is None or voorschot_val <= 0:
            ws_sheet.cell(r, 4).value = "—"
            ws_sheet.cell(r, 4).alignment = Alignment(horizontal="right")
        else:
            val(ws_sheet.cell(r, 4), voorschot_val, "€ #,##0")

    # Auteur
    if ti.get("auteur_royalty_staffel"):
        # Toon eerste trede; meerdere treden als sub-info
        staffel = ti["auteur_royalty_staffel"]
        first_pct = staffel[0].get("percentage", 0) if staffel else 0
        deal_row("Auteur", "Royalty-staffel", first_pct, ti.get("auteur_voorschot", 0))
        r += 1
        if len(staffel) > 1:
            for trede in staffel:
                ws_sheet.cell(r, 1).value = f"   tot {trede.get('tot_exemplaren', 0):,} ex"
                ws_sheet.cell(r, 1).font = Font(size=9, italic=True, color="FF666666")
                val(ws_sheet.cell(r, 3), trede.get("percentage", 0), "0.0%")
                r += 1
    elif ti.get("auteur_winstdeling_pct", 0) > 0:
        deal_row("Auteur", "Winstdeling", ti.get("auteur_winstdeling_pct", 0), None)
        r += 1
    else:
        deal_row("Auteur", "—", None, None)
        r += 1

    # Agent, Vertaler, Illustrator
    for partij_key, partij_naam in [("agent", "Agent"), ("vertaler", "Vertaler"), ("illustrator", "Illustrator")]:
        staffel = ti.get(f"{partij_key}_staffel") or []
        winst = ti.get(f"{partij_key}_winstdeling_pct", 0)
        vast = ti.get(f"{partij_key}_pct", 0)
        voorschot = ti.get(f"{partij_key}_voorschot", 0)
        # Skip helemaal als geen enkele waarde
        if not staffel and winst == 0 and vast == 0 and (voorschot or 0) == 0:
            continue
        if staffel:
            first_pct = staffel[0].get("percentage", 0)
            deal_row(partij_naam, "Royalty-staffel", first_pct, voorschot)
        elif winst > 0:
            deal_row(partij_naam, "Winstdeling", winst, None)
        elif vast > 0:
            deal_row(partij_naam, "Vast %", vast, voorschot)
        else:
            # Alleen een achtergebleven voorschot zonder deal — laat 't zien als "geen deal"
            deal_row(partij_naam, "(geen deal)", None, voorschot)
        r += 1

    # Extra derden
    for ed in (ti.get("extra_derden") or []):
        naam = ed.get("naam") or "Extra persoon"
        type_str = ed.get("type", "royalty")
        if type_str == "winstdeling":
            deal_row(naam, "Winstdeling", ed.get("percentage", 0), None)
        elif ed.get("staffel"):
            first_pct = ed["staffel"][0].get("percentage", 0)
            deal_row(naam, "Royalty-staffel", first_pct, ed.get("voorschot", 0))
        else:
            deal_row(naam, "Vast %", ed.get("percentage", 0), ed.get("voorschot", 0))
        r += 1

    # Partner
    if ti.get("heeft_partner"):
        partner_naam = ti.get("partner_naam") or "Partner"
        deal_row(partner_naam, "Partnership (informatief)", ti.get("partner_winstdeling_pct", 0.5), None)
        r += 1
        ws_sheet.merge_cells(f"A{r}:F{r}")
        ws_sheet[f"A{r}"].value = "Partner-winstdeling staat buiten de titel-marge (geregeld via productiehuis-overhead)."
        ws_sheet[f"A{r}"].font = Font(size=8, italic=True, color="FF999999")
        r += 1

    # ──────────────────────────────────────────────────────────────────
    #  TAB 2 — RESULTAAT (outputs)
    # ──────────────────────────────────────────────────────────────────
    ws_res = wb.create_sheet("Resultaat")
    set_widths(ws_res, 36, 16, 16, 16, 16, 16, 16, 16)
    r = 1

    # ── HEADER ──
    ws_res.merge_cells(f"A{r}:H{r}")
    h1(ws_res[f"A{r}"], f"Resultaat — {ti.get('titel', '?')}")
    ws_res.row_dimensions[r].height = 24
    r += 2

    # Bereken aggregaten voor headline
    druk0 = calc["drukken"][0] if calc.get("drukken") else {}
    # Gebruik gewogen_marge_pct_totaal (over ALLE drukken), niet alleen druk0
    gewogen_marge = calc.get("gewogen_marge_pct_totaal", druk0.get("gewogen_marge_pct", 0) if druk0 else 0)
    # Gewogen netto winst en omzet per ex: som euro / totaal oplage
    totaal_oplage = sum(d.get("oplage", 0) for d in drukken_cfg)
    if totaal_oplage > 0:
        gewogen_winst = sum(d["gewogen_netto_winst"] * d["oplage"] for d in calc["drukken"]) / totaal_oplage
        gewogen_omzet = sum(d["gewogen_netto_omzet"] * d["oplage"] for d in calc["drukken"]) / totaal_oplage
    else:
        gewogen_winst = druk0.get("gewogen_netto_winst", 0) if druk0 else 0
        gewogen_omzet = druk0.get("gewogen_netto_omzet", 0) if druk0 else 0

    # ── HEADLINE ──
    ws_res.merge_cells(f"A{r}:H{r}")
    h2(ws_res[f"A{r}"], "GEWOGEN MARGE")
    r += 1
    label(ws_res[f"A{r}"], "Marge", bold=True)
    val(ws_res[f"B{r}"], gewogen_marge, "0.0%", bold=True,
        color=GROEN if gewogen_marge >= 0.35 else ("FFE65100" if gewogen_marge >= 0 else "FFC62828"))
    label(ws_res[f"D{r}"], "Netto winst per ex", bold=True)
    val(ws_res[f"E{r}"], gewogen_winst, "€ #,##0.00", bold=True)
    r += 1
    label(ws_res[f"A{r}"], "Streefmarge")
    val(ws_res[f"B{r}"], 0.35, "0.0%")
    label(ws_res[f"D{r}"], "Netto omzet per ex")
    val(ws_res[f"E{r}"], gewogen_omzet, "€ #,##0.00")
    r += 1
    label(ws_res[f"A{r}"], "Totaal exemplaren (alle drukken)")
    val(ws_res[f"B{r}"], totaal_oplage, "#,##0")
    r += 2

    # ── MARGE PER KANAAL — WATERFALL ──
    if druk0:
        ws_res.merge_cells(f"A{r}:H{r}")
        h2(ws_res[f"A{r}"], f"MARGE PER KANAAL — 1e druk ({druk0.get('oplage', 0):,} ex.)")
        r += 1

        rt = druk0.get("retail", {})
        ws_d = druk0.get("webshop", {})
        b2b = druk0.get("b2b", {})

        def gewogen(field):
            return ws_d.get(field, 0)*verd_ws + rt.get(field, 0)*verd_rt + b2b.get(field, 0)*verd_b2b

        def gewogen_extra_total():
            total = 0.0
            for kanaal_data, w in [(rt, verd_rt), (ws_d, verd_ws), (b2b, verd_b2b)]:
                for ed in (kanaal_data.get("extra_derden_per_naam") or []):
                    total += (ed.get("bedrag", 0) or 0) * w
            return total

        # Kolom-headers
        for ci, hdr in enumerate(["Kostenregel", "Retail/CB", "Webshop", "B2B", "Gewogen"], 1):
            th(ws_res.cell(r, ci), hdr)
            if ci == 1:
                ws_res.cell(r, ci).alignment = Alignment(horizontal="left", indent=1)
        r += 1

        # Verzamel mode-info voor agent/vertaler/illustrator
        def _is_winstdeling(partij: str) -> bool:
            return ti.get(f"{partij}_winstdeling_pct", 0) > 0

        # Bouw rij-definitie. Type: 'item' / 'subtotal' / 'info' / 'extra_per_kanaal'
        # Voor extra_per_kanaal: render per kanaal de eigen list
        waterfall_rows: list[dict] = []

        def add(label_txt, field, type_="item", sign=-1, only_if=None):
            waterfall_rows.append({"label": label_txt, "field": field, "type": type_, "sign": sign, "only_if": only_if})

        # Omzet sectie
        waterfall_rows.append({"label": "Verkoopprijs ex BTW", "field": "verkoopprijs_ex_btw", "type": "item", "sign": 1})
        waterfall_rows.append({"label": "Korting", "field": "korting_bedrag", "type": "item", "sign": -1, "only_if": lambda data: data.get("korting_bedrag", 0) > 0})
        waterfall_rows.append({"label": "Netto omzet", "field": "netto_omzet", "type": "subtotal", "sign": 1})
        # Operationele kosten
        waterfall_rows.append({"label": "Drukkosten /ex", "field": "drukkosten", "type": "item", "sign": -1})
        waterfall_rows.append({"label": "Kostenposten /ex", "field": "kosten_per_ex", "type": "item", "sign": -1})
        waterfall_rows.append({"label": "Fulfillment", "field": "fulfillment", "type": "item", "sign": -1})
        waterfall_rows.append({"label": "Distributie CB", "field": "distributie_cb", "type": "item", "sign": -1})
        waterfall_rows.append({"label": "B2B porto", "field": "b2b_porto", "type": "item", "sign": -1})
        waterfall_rows.append({"label": "Transactiekosten", "field": "transactiekosten", "type": "item", "sign": -1})
        waterfall_rows.append({"label": "CAC", "field": "cac", "type": "item", "sign": -1})
        # Royalty-derden (boven brutowinst, alleen als royalty-mode)
        waterfall_rows.append({"label": "Auteur royalty", "field": "auteur_royalty", "type": "item", "sign": -1})
        if not _is_winstdeling("vertaler"):
            waterfall_rows.append({"label": "Vertaler", "field": "vertaler", "type": "item", "sign": -1})
        if not _is_winstdeling("illustrator"):
            waterfall_rows.append({"label": "Illustrator", "field": "illustrator", "type": "item", "sign": -1})
        if not _is_winstdeling("agent"):
            waterfall_rows.append({"label": "Agent", "field": "agent", "type": "item", "sign": -1})
        # Extra royalty-derden (boven brutowinst): hier renderen we ze als één samengevoegde regel per persoon
        # via een speciale 'extra_royalty' type
        waterfall_rows.append({"label": "__extra_royalty__", "field": None, "type": "extra_royalty", "sign": -1})
        waterfall_rows.append({"label": "Overige kosten", "field": "overige_kosten", "type": "item", "sign": -1})
        waterfall_rows.append({"label": "Brutowinst", "field": "brutowinst", "type": "subtotal", "sign": 1})
        # Winstdeling-derden (onder brutowinst)
        waterfall_rows.append({"label": "Auteur winstdeling", "field": "auteur_winstdeling", "type": "item", "sign": -1})
        if _is_winstdeling("vertaler"):
            waterfall_rows.append({"label": "Vertaler (winstdeling)", "field": "vertaler", "type": "item", "sign": -1})
        if _is_winstdeling("illustrator"):
            waterfall_rows.append({"label": "Illustrator (winstdeling)", "field": "illustrator", "type": "item", "sign": -1})
        if _is_winstdeling("agent"):
            waterfall_rows.append({"label": "Agent (winstdeling)", "field": "agent", "type": "item", "sign": -1})
        waterfall_rows.append({"label": "__extra_winstdeling__", "field": None, "type": "extra_winstdeling", "sign": -1})
        waterfall_rows.append({"label": "Netto winst Maven", "field": "netto_winst_maven", "type": "subtotal", "sign": 1})
        waterfall_rows.append({"label": "Marge %", "field": "marge_pct", "type": "marge", "sign": 1})
        if ti.get("heeft_partner"):
            waterfall_rows.append({"label": "Partner-winstdeling (informatief)", "field": "partner_winstdeling", "type": "info", "sign": -1})

        # Helper voor per-kanaal extra derden bedragen
        def extra_per_kanaal(kanaal_data, want_type: str) -> dict[str, float]:
            out: dict[str, float] = {}
            for ed in (kanaal_data.get("extra_derden_per_naam") or []):
                if ed.get("type") == want_type:
                    out[ed.get("naam", "Extra")] = (ed.get("bedrag", 0) or 0)
            return out

        # Render de waterfall
        for wf in waterfall_rows:
            if wf["type"] in ("extra_royalty", "extra_winstdeling"):
                want = "royalty" if wf["type"] == "extra_royalty" else "winstdeling"
                # Verzamel alle unieke namen over de kanalen
                names: set[str] = set()
                for kanaal_data in (rt, ws_d, b2b):
                    for ed in (kanaal_data.get("extra_derden_per_naam") or []):
                        if ed.get("type") == want:
                            names.add(ed.get("naam", "Extra"))
                for nm in sorted(names):
                    suffix = " (winstdeling)" if want == "winstdeling" else ""
                    ws_res.cell(r, 1).value = f"  {nm}{suffix}"
                    ws_res.cell(r, 1).font = Font(size=9)
                    ws_res.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
                    for ci, (kanaal_data, _) in enumerate([(rt, verd_rt), (ws_d, verd_ws), (b2b, verd_b2b)], 2):
                        ed_map = extra_per_kanaal(kanaal_data, want)
                        v = -(ed_map.get(nm, 0))
                        c = ws_res.cell(r, ci)
                        val(c, v, "€ #,##0.00")
                    # Gewogen
                    gw = -(extra_per_kanaal(rt, want).get(nm, 0)*verd_rt
                         + extra_per_kanaal(ws_d, want).get(nm, 0)*verd_ws
                         + extra_per_kanaal(b2b, want).get(nm, 0)*verd_b2b)
                    val(ws_res.cell(r, 5), gw, "€ #,##0.00")
                    r += 1
                continue

            # Filter: only_if
            if wf.get("only_if"):
                # gebruik gewogen data voor filter
                gewogen_data = {f: gewogen(f) for f in ["korting_bedrag"]}
                if not wf["only_if"](gewogen_data):
                    continue

            is_subtotal = wf["type"] == "subtotal"
            is_info = wf["type"] == "info"
            is_marge = wf["type"] == "marge"

            # Label
            c_a = ws_res.cell(r, 1)
            c_a.value = f"  {wf['label']}" if not (is_subtotal or is_marge) else wf["label"]
            c_a.font = Font(
                size=9,
                bold=is_subtotal or is_marge,
                italic=is_info,
                color="FF999999" if is_info else "FF000000",
            )
            c_a.alignment = Alignment(horizontal="left", indent=1 if not (is_subtotal or is_marge) else 0)

            # Per kanaal
            for ci, (kanaal_data, _) in enumerate([(rt, verd_rt), (ws_d, verd_ws), (b2b, verd_b2b)], 2):
                v = kanaal_data.get(wf["field"], 0) * wf["sign"]
                c = ws_res.cell(r, ci)
                if is_marge:
                    # Marge t.o.v. netto omzet — gebruik direct uit kanaal
                    val(c, kanaal_data.get("marge_pct", 0), "0.0%",
                        bold=True,
                        color=GROEN if kanaal_data.get("marge_pct", 0) >= 0.35 else ("FFE65100" if kanaal_data.get("marge_pct", 0) >= 0 else "FFC62828"))
                else:
                    val(c, v, "€ #,##0.00",
                        bold=is_subtotal,
                        color="FF999999" if is_info else "FF000000")
                if is_subtotal:
                    c.fill = PatternFill("solid", fgColor=LICHT_GRIJS)

            # Gewogen
            gew_cell = ws_res.cell(r, 5)
            if is_marge:
                gw_marge = gewogen("netto_winst_maven") / gewogen("netto_omzet") if gewogen("netto_omzet") > 0 else 0
                val(gew_cell, gw_marge, "0.0%",
                    bold=True,
                    color=GROEN if gw_marge >= 0.35 else ("FFE65100" if gw_marge >= 0 else "FFC62828"))
            else:
                gw_v = gewogen(wf["field"]) * wf["sign"]
                val(gew_cell, gw_v, "€ #,##0.00",
                    bold=is_subtotal,
                    color="FF999999" if is_info else "FF000000")
            if is_subtotal:
                gew_cell.fill = PatternFill("solid", fgColor=LICHT_GRIJS)
                c_a.fill = PatternFill("solid", fgColor=LICHT_GRIJS)
            r += 1
        r += 1

    # ── OPLAGE SIMULATIE ──
    ws_res.merge_cells(f"A{r}:H{r}")
    h2(ws_res[f"A{r}"], "OPLAGE SIMULATIE")
    r += 1

    sim_headers = ["Oplage", "Netto omzet", "Drukkosten", "Eenmalige kosten", "Dealkosten", "Netto resultaat", "Marge %"]
    for ci, hdr in enumerate(sim_headers, 1):
        th(ws_res.cell(r, ci), hdr)
        ws_res.cell(r, ci).alignment = Alignment(horizontal="right" if ci > 1 else "left")
    r += 1

    # Bereken sim_at met dealkosten samengevoegd (zonder aparte voorschot-kolom)
    def _heeft_royalty_deal_x(prefix: str) -> bool:
        return ti.get(f"{prefix}_pct", 0) > 0 or len(ti.get(f"{prefix}_staffel") or []) > 0

    auteur_vs = ti.get("auteur_voorschot", 0) if len(ti.get("auteur_royalty_staffel") or []) > 0 else 0
    agent_vs = ti.get("agent_voorschot", 0) if _heeft_royalty_deal_x("agent") else 0
    vertaler_vs = ti.get("vertaler_voorschot", 0) if _heeft_royalty_deal_x("vertaler") else 0
    illustrator_vs = ti.get("illustrator_voorschot", 0) if _heeft_royalty_deal_x("illustrator") else 0
    extra_vs = sum(
        d.get("voorschot", 0) for d in ti.get("extra_derden", [])
        if d.get("type") == "royalty" and (
            d.get("percentage", 0) > 0 or len(d.get("staffel") or []) > 0
        )
    )
    totaal_vs = auteur_vs + agent_vs + vertaler_vs + illustrator_vs + extra_vs

    ws_k = druk0.get("webshop", {})
    rt_k = druk0.get("retail", {})
    b2b_k_d = druk0.get("b2b", {})

    def gew_sim(field):
        return ws_k.get(field, 0)*verd_ws + rt_k.get(field, 0)*verd_rt + b2b_k_d.get(field, 0)*verd_b2b

    netto_omzet_pex = gew_sim("netto_omzet")
    netto_winst_pex = gew_sim("netto_winst_maven")
    kosten_pex_sim = gew_sim("kosten_per_ex")
    druk_pex = gew_sim("drukkosten")
    royalty_pex = gew_sim("auteur_royalty")
    agent_pex = gew_sim("agent")
    vertaler_pex = gew_sim("vertaler")
    illustrator_pex = gew_sim("illustrator")

    var_w = netto_winst_pex + kosten_pex_sim
    pure_v = var_w + druk_pex

    adj_pex = pure_v
    if auteur_vs > 0:
        adj_pex += royalty_pex
    if agent_vs > 0 and agent_pex > 0:
        adj_pex += agent_pex
    if vertaler_vs > 0 and vertaler_pex > 0:
        adj_pex += vertaler_pex
    if illustrator_vs > 0 and illustrator_pex > 0:
        adj_pex += illustrator_pex

    totaal_eenmalig_sim = druk0.get("kosten_totaal", 0)

    def sim_at_v2(vol):
        # Drukkosten per druk capped op oplage
        dk = 0
        rem = vol
        for d in sorted(drukken_cfg, key=lambda x: x.get("druknummer", 1)):
            dv = min(rem, d.get("oplage", 0))
            dk += dv * d.get("drukkosten_per_ex", 0)
            rem -= dv
            if rem <= 0:
                break
        if rem > 0 and drukken_cfg:
            dk += rem * drukken_cfg[-1].get("drukkosten_per_ex", 1.20)

        # Dealkosten: voorschot (vast) + royalty boven voorschot
        dealkosten = 0.0
        if auteur_vs > 0:
            auteur_earned = vol * royalty_pex
            dealkosten += max(auteur_vs, auteur_earned) if royalty_pex > 0 else auteur_vs
        if agent_vs > 0:
            agent_earned = vol * agent_pex
            dealkosten += max(agent_vs, agent_earned) if agent_pex > 0 else agent_vs
        if vertaler_vs > 0:
            vertaler_earned = vol * vertaler_pex
            dealkosten += max(vertaler_vs, vertaler_earned) if vertaler_pex > 0 else vertaler_vs
        if illustrator_vs > 0:
            illustrator_earned = vol * illustrator_pex
            dealkosten += max(illustrator_vs, vol * illustrator_pex) if illustrator_pex > 0 else illustrator_vs
        dealkosten += extra_vs

        omzet = vol * netto_omzet_pex
        net = vol * adj_pex - dk - totaal_eenmalig_sim - dealkosten
        marge = net / omzet if omzet > 0 else 0

        # Voorschot ingelopen check
        total_earned = (
            (vol * royalty_pex if auteur_vs > 0 else 0)
            + (vol * agent_pex if agent_vs > 0 else 0)
            + (vol * vertaler_pex if vertaler_vs > 0 else 0)
            + (vol * illustrator_pex if illustrator_vs > 0 else 0)
        )
        voorschot_in = total_earned >= totaal_vs if totaal_vs > 0 else None
        return {"vol": vol, "omzet": omzet, "drukkosten": dk, "eenmalig": totaal_eenmalig_sim,
                "dealkosten": dealkosten, "net": net, "marge": marge, "voorschot_in": voorschot_in}

    def find_be_v2():
        if sim_at_v2(200000)["net"] < 0:
            return None
        if sim_at_v2(1)["net"] >= 0:
            pl_be = 0
        else:
            lo, hi = 1, 200000
            for _ in range(50):
                mid = (lo + hi) // 2
                if sim_at_v2(mid)["net"] < 0:
                    lo = mid
                else:
                    hi = mid
                if hi - lo <= 10:
                    break
            pl_be = hi
        return ((pl_be + 49) // 50) * 50

    totaal_oplage_sim = sum(d.get("oplage", 0) for d in drukken_cfg)
    be_vol = find_be_v2()

    sim_vols = sorted(set(filter(None, [
        be_vol,
        totaal_oplage_sim,
        totaal_oplage_sim + 2500,
        totaal_oplage_sim + 5000,
        totaal_oplage_sim + 10000,
        totaal_oplage_sim + 20000,
    ])))[:7]

    for sv in sim_vols:
        sd = sim_at_v2(sv)
        is_be = (be_vol is not None and sv == be_vol)
        bg = LICHTGROEN if is_be else WIT

        # Oplage label
        label_text = f"{sv:,}".replace(",", ".") + ("  ← break-even" if is_be else "")
        c0 = ws_res.cell(r, 1)
        c0.value = label_text
        c0.font = Font(size=9, bold=is_be)
        c0.fill = PatternFill("solid", fgColor=bg)

        def sc(ci, v, fmt):
            c = ws_res.cell(r, ci)
            c.value = v
            c.number_format = fmt
            c.font = Font(size=9, bold=is_be)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="right")

        sc(2, sd["omzet"], "€ #,##0")
        sc(3, sd["drukkosten"], "€ #,##0")
        sc(4, sd["eenmalig"], "€ #,##0")
        sc(5, sd["dealkosten"], "€ #,##0")
        sc(6, sd["net"], "€ #,##0")
        # Marge met kleur
        marg_c = ws_res.cell(r, 7)
        marg_c.value = sd["marge"]
        marg_c.number_format = "0.0%"
        marg_c.font = Font(size=9, bold=is_be,
                           color=GROEN if sd["marge"] >= 0.35 else ("FFE65100" if sd["marge"] >= 0 else "FFC62828"))
        marg_c.fill = PatternFill("solid", fgColor=bg)
        marg_c.alignment = Alignment(horizontal="right")
        r += 1

    # Toelichting onderaan oplage-simulatie
    r += 1
    ws_res.merge_cells(f"A{r}:G{r}")
    note_cell = ws_res[f"A{r}"]
    note_cell.value = (
        "Dealkosten = voorschotten + royalty's en commissies aan alle partijen, inclusief winstdeling."
    )
    note_cell.font = Font(size=8, italic=True, color="FF666666")
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    r += 2

    # ── VOORSCHOT INGELOPEN ──
    # Toon alleen als er een actief voorschot is om in te lopen
    if totaal_vs > 0:
        ws_res.merge_cells(f"A{r}:G{r}")
        h2(ws_res[f"A{r}"], "VOORSCHOT INGELOPEN")
        r += 1

        # Bereken earn-out volume (cumulatieve royalty's ≥ voorschot)
        total_recoup_per_ex = (
            (royalty_pex if auteur_vs > 0 else 0)
            + (agent_pex if agent_vs > 0 else 0)
            + (vertaler_pex if vertaler_vs > 0 else 0)
            + (illustrator_pex if illustrator_vs > 0 else 0)
        )
        earn_out_vol = None
        if total_recoup_per_ex > 0:
            earn_out_vol = int(totaal_vs / total_recoup_per_ex) + 1
            # Rond op nearest 50
            earn_out_vol = ((earn_out_vol + 49) // 50) * 50

        # Samenvatting
        label(ws_res[f"A{r}"], "Totaal voorschot", bold=True)
        val(ws_res[f"B{r}"], totaal_vs, "€ #,##0", bold=True)
        r += 1
        label(ws_res[f"A{r}"], "Royalty / commissie per ex")
        val(ws_res[f"B{r}"], total_recoup_per_ex, "€ #,##0.00")
        r += 1
        label(ws_res[f"A{r}"], "Ingelopen bij oplage", bold=True)
        if earn_out_vol is None:
            ws_res[f"B{r}"].value = "n.v.t. (geen royalty-stroom)"
            ws_res[f"B{r}"].font = Font(size=10, italic=True, color="FF999999")
            ws_res[f"B{r}"].alignment = Alignment(horizontal="right")
        else:
            val(ws_res[f"B{r}"], earn_out_vol, "#,##0", bold=True,
                color=GROEN)
            ws_res.cell(r, 3).value = "ex"
            ws_res.cell(r, 3).font = Font(size=9, color="FF999999")
        r += 1

        # Per-partij breakdown (alleen actieve partijen)
        breakdowns = []
        if auteur_vs > 0 and royalty_pex > 0:
            breakdowns.append(("Auteur", auteur_vs, royalty_pex))
        if agent_vs > 0 and agent_pex > 0:
            breakdowns.append(("Agent", agent_vs, agent_pex))
        if vertaler_vs > 0 and vertaler_pex > 0:
            breakdowns.append(("Vertaler", vertaler_vs, vertaler_pex))
        if illustrator_vs > 0 and illustrator_pex > 0:
            breakdowns.append(("Illustrator", illustrator_vs, illustrator_pex))

        if len(breakdowns) > 1:
            # Tabel-header
            r += 1
            for ci, hdr in enumerate(["Partij", "Voorschot", "Royalty / ex", "Ingelopen bij"], 1):
                th(ws_res.cell(r, ci), hdr)
                ws_res.cell(r, ci).alignment = Alignment(horizontal="right" if ci > 1 else "left")
            r += 1
            for naam, vs, royalty in breakdowns:
                ws_res.cell(r, 1).value = naam
                ws_res.cell(r, 1).font = Font(size=10)
                val(ws_res.cell(r, 2), vs, "€ #,##0")
                val(ws_res.cell(r, 3), royalty, "€ #,##0.00")
                partij_earn = ((int(vs / royalty) + 49) // 50) * 50 if royalty > 0 else None
                if partij_earn is not None:
                    val(ws_res.cell(r, 4), partij_earn, "#,##0")
                r += 1

        # Status per oplage-punt
        r += 1
        for ci, hdr in enumerate(["Oplage", "Royalty's verdiend", "Status"], 1):
            th(ws_res.cell(r, ci), hdr)
            ws_res.cell(r, ci).alignment = Alignment(horizontal="right" if ci > 1 else "left")
        r += 1

        for sv in sim_vols:
            total_earned = sv * total_recoup_per_ex
            ingelopen = total_earned >= totaal_vs

            ws_res.cell(r, 1).value = f"{sv:,}".replace(",", ".")
            ws_res.cell(r, 1).font = Font(size=9)
            val(ws_res.cell(r, 2), total_earned, "€ #,##0")
            status_c = ws_res.cell(r, 3)
            if ingelopen:
                status_c.value = "✓ Ingelopen"
                status_c.font = Font(size=9, bold=True, color=GROEN)
            else:
                pct_done = total_earned / totaal_vs if totaal_vs > 0 else 0
                status_c.value = f"Nog open: {(totaal_vs - total_earned):,.0f}".replace(",", ".") + f"  ({pct_done*100:.0f}%)"
                status_c.font = Font(size=9, color="FFE65100")
            status_c.alignment = Alignment(horizontal="right")
            r += 1

    # Output
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    titel_slug = (ti.get("titel", "calculatie") or "calculatie").replace(" ", "_")[:30]
    filename = f"calculatie_{titel_slug}.xlsx"

    from flask import send_file
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── PDF Export ──

@bp.route("/api/export/pdf", methods=["POST"])
def export_pdf():
    """Exporteer calculatie-resultaat als PDF (A4 staand).

    Bevat: gewogen marge headline, marge-per-kanaal waterfall (1e druk),
    en oplage-simulatietabel. Opmaak sluit aan op Excel-export.
    """
    from io import BytesIO
    from datetime import date as _date
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

    data = request.get_json()
    if not data:
        abort(400, description="Geen data meegegeven")

    ti = data.get("titel_input", {})
    verd_ws = data.get("verdeling_webshop", 0.10)
    verd_rt = data.get("verdeling_retail", 0.85)
    verd_b2b = data.get("verdeling_b2b", 0.05)

    try:
        calc = run_calculation(data)
    except Exception as e:
        abort(400, description=f"Berekening mislukt: {e}")

    druk0 = calc["drukken"][0] if calc.get("drukken") else {}
    drukken_cfg = ti.get("drukken", [])
    totaal_oplage = sum(d.get("oplage", 0) for d in drukken_cfg)

    # ── Kleuren (gelijk aan Excel-export) ──
    C_DONKER = colors.HexColor("#37474F")   # header achtergrond
    C_WIT = colors.white
    C_GROEN_DONKER = colors.HexColor("#1B5E20")
    C_GROEN_LICHT = colors.HexColor("#E8F5E9")
    C_GRIJS_LICHT = colors.HexColor("#EEEEEE")
    C_ORANJE = colors.HexColor("#E65100")
    C_ROOD = colors.HexColor("#C62828")

    def marge_kleur(pct: float):
        if pct >= 0.35:
            return C_GROEN_DONKER
        if pct >= 0:
            return C_ORANJE
        return C_ROOD

    # ── Stijlen ──
    FONT = "Helvetica"
    FONT_B = "Helvetica-Bold"

    def ps(name, **kwargs):
        defaults = dict(fontName=FONT, fontSize=9, leading=12)
        defaults.update(kwargs)
        return ParagraphStyle(name, **defaults)

    sNorm = ps("norm")
    sRight = ps("right", alignment=TA_RIGHT)
    sBold = ps("bold", fontName=FONT_B)
    sBoldRight = ps("boldright", fontName=FONT_B, alignment=TA_RIGHT)
    sHeader = ps("header", fontName=FONT_B, fontSize=13, textColor=C_WIT)
    sSection = ps("section", fontName=FONT_B, fontSize=9, textColor=C_GROEN_DONKER)
    sMeta = ps("meta", fontSize=8, textColor=colors.HexColor("#666666"))

    buf = BytesIO()
    W, H = A4
    margin = 20 * mm
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=margin, rightMargin=margin,
        topMargin=margin, bottomMargin=margin,
    )

    story = []

    def spacer(h=3):
        story.append(Spacer(1, h * mm))

    def section_header(tekst):
        usable = W - 2 * margin
        tbl = Table([[Paragraph(tekst, sSection)]], colWidths=[usable])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), C_GROEN_LICHT),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(tbl)
        spacer(2)

    # ── 1. DOCUMENT HEADER ──
    usable = W - 2 * margin
    titel_str = ti.get("titel", "—")
    auteur_str = ti.get("auteur", "") or "—"
    isbn_str = ti.get("isbn", "")
    meta_parts = [auteur_str]
    if isbn_str:
        meta_parts.append(f"ISBN {isbn_str}")
    meta_parts.append(f"Gegenereerd {_date.today().strftime('%d-%m-%Y')}")

    hdr_tbl = Table(
        [[Paragraph(f"Calculatie — {titel_str}", sHeader)],
         [Paragraph("  ·  ".join(meta_parts), sMeta)]],
        colWidths=[usable],
    )
    hdr_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), C_DONKER),
        ("TOPPADDING", (0, 0), (0, 0), 8),
        ("BOTTOMPADDING", (0, 0), (0, 0), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 1), (0, 1), 4),
        ("BOTTOMPADDING", (0, 1), (0, 1), 4),
    ]))
    story.append(hdr_tbl)
    spacer(4)

    # ── 2. GEWOGEN MARGE ──
    section_header("GEWOGEN MARGE")

    gewogen_marge = calc.get("gewogen_marge_pct_totaal", 0)
    if totaal_oplage > 0 and calc.get("drukken"):
        gw = sum(d["gewogen_netto_winst"] * d["oplage"] for d in calc["drukken"]) / totaal_oplage
        go = sum(d["gewogen_netto_omzet"] * d["oplage"] for d in calc["drukken"]) / totaal_oplage
    else:
        gw = druk0.get("gewogen_netto_winst", 0) if druk0 else 0
        go = druk0.get("gewogen_netto_omzet", 0) if druk0 else 0

    marge_c = marge_kleur(gewogen_marge)
    col1 = usable * 0.5
    col2 = usable * 0.5
    headline_data = [
        [Paragraph("Gewogen marge", sBold),
         Paragraph("Streefmarge", sNorm)],
        [Paragraph(f'<font color="{marge_c.hexval()}" size="20"><b>{gewogen_marge*100:.1f}%</b></font>', ParagraphStyle("m", alignment=TA_LEFT)),
         Paragraph("35,0%", ps("s35", fontSize=14, textColor=colors.HexColor("#999999")))],
        [Paragraph(f"Netto winst per ex: <b>€ {gw:.2f}</b>  ·  Netto omzet per ex: <b>€ {go:.2f}</b>  ·  Totaal exemplaren: <b>{totaal_oplage:,}</b>".replace(",", "."), ps("sub", fontSize=8, textColor=colors.HexColor("#555555"))),
         Paragraph("")],
    ]
    hl_tbl = Table(headline_data, colWidths=[col1, col2])
    hl_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("SPAN", (0, 2), (1, 2)),
    ]))
    story.append(hl_tbl)
    spacer(5)

    # ── 3. MARGE PER KANAAL — WATERFALL ──
    if druk0:
        oplage_d0 = druk0.get("oplage", 0)
        section_header(f"MARGE PER KANAAL — 1e druk ({oplage_d0:,} ex.)".replace(",", "."))

        rt = druk0.get("retail", {})
        ws_d = druk0.get("webshop", {})
        b2b_k = druk0.get("b2b", {})

        def gew(field):
            return ws_d.get(field, 0)*verd_ws + rt.get(field, 0)*verd_rt + b2b_k.get(field, 0)*verd_b2b

        def _is_wd(partij):
            return ti.get(f"{partij}_winstdeling_pct", 0) > 0

        def fmt_eur(v, sign=1):
            if v == 0:
                return "—"
            return f"€ {v*sign:+.2f}".replace("+", "").replace(",", ".")

        def fmt_pct(v):
            return f"{v*100:.1f}%"

        # Kolom-breedtes: label 150pt + 4×(usable-150)/4
        lbl_w = 155
        val_w = (usable - lbl_w) / 4

        wf_headers = ["Kostenregel", "Retail/CB", "Webshop", "B2B", "Gewogen"]
        wf_rows = [wf_headers]
        row_styles = []
        row_idx = 1  # header is row 0

        def add_row(label_txt, rt_v, ws_v, b2b_v, gew_v, bold=False, bg=None, is_marge=False):
            nonlocal row_idx
            sL = sBold if bold else sNorm
            sR = sBoldRight if bold else sRight

            if is_marge:
                rt_p = Paragraph(f'<font color="{marge_kleur(rt_v).hexval()}"><b>{fmt_pct(rt_v)}</b></font>', sRight)
                ws_p = Paragraph(f'<font color="{marge_kleur(ws_v).hexval()}"><b>{fmt_pct(ws_v)}</b></font>', sRight)
                b2b_p = Paragraph(f'<font color="{marge_kleur(b2b_v).hexval()}"><b>{fmt_pct(b2b_v)}</b></font>', sRight)
                gw_pct = gew("netto_winst_maven") / gew("netto_omzet") if gew("netto_omzet") > 0 else 0
                gew_p = Paragraph(f'<font color="{marge_kleur(gw_pct).hexval()}"><b>{fmt_pct(gw_pct)}</b></font>', sRight)
                wf_rows.append([Paragraph(label_txt, sL), rt_p, ws_p, b2b_p, gew_p])
            else:
                wf_rows.append([
                    Paragraph(label_txt, sL),
                    Paragraph(fmt_eur(rt_v), sR),
                    Paragraph(fmt_eur(ws_v), sR),
                    Paragraph(fmt_eur(b2b_v), sR),
                    Paragraph(fmt_eur(gew_v), sR),
                ])

            if bg:
                row_styles.append(("BACKGROUND", (0, row_idx), (-1, row_idx), bg))
            if bold:
                row_styles.append(("FONTNAME", (0, row_idx), (-1, row_idx), FONT_B))
            row_idx += 1

        def add_sep():
            nonlocal row_idx
            wf_rows.append(["", "", "", "", ""])
            row_styles.append(("LINEBELOW", (0, row_idx - 1), (-1, row_idx - 1), 0.3, C_GRIJS_LICHT))

        # Omzet
        add_row("Verkoopprijs ex BTW",
                rt.get("verkoopprijs_ex_btw", 0), ws_d.get("verkoopprijs_ex_btw", 0), b2b_k.get("verkoopprijs_ex_btw", 0), gew("verkoopprijs_ex_btw"))
        if gew("korting_bedrag") > 0:
            add_row("  Korting",
                    -rt.get("korting_bedrag", 0), -ws_d.get("korting_bedrag", 0), -b2b_k.get("korting_bedrag", 0), -gew("korting_bedrag"))
        add_row("Netto omzet",
                rt.get("netto_omzet", 0), ws_d.get("netto_omzet", 0), b2b_k.get("netto_omzet", 0), gew("netto_omzet"),
                bold=True, bg=C_GRIJS_LICHT)

        # Operationele kosten
        def add_if_nonzero(lbl, field):
            if gew(field) > 0 or any(k.get(field, 0) for k in (rt, ws_d, b2b_k)):
                add_row(f"  {lbl}",
                        -rt.get(field, 0), -ws_d.get(field, 0), -b2b_k.get(field, 0), -gew(field))

        add_if_nonzero("Drukkosten /ex", "drukkosten")
        add_if_nonzero("Kostenposten /ex", "kosten_per_ex")
        add_if_nonzero("Fulfillment", "fulfillment")
        add_if_nonzero("Distributie CB", "distributie_cb")
        add_if_nonzero("B2B porto", "b2b_porto")
        add_if_nonzero("Transactiekosten", "transactiekosten")
        add_if_nonzero("CAC", "cac")

        # Royalty-derden boven brutowinst
        if not _is_wd("auteur") and gew("auteur_royalty") > 0:
            add_if_nonzero("Auteur royalty", "auteur_royalty")
        if not _is_wd("vertaler") and gew("vertaler") > 0:
            add_if_nonzero("Vertaler", "vertaler")
        if not _is_wd("illustrator") and gew("illustrator") > 0:
            add_if_nonzero("Illustrator", "illustrator")
        if not _is_wd("agent") and gew("agent") > 0:
            add_if_nonzero("Agent", "agent")

        # Extra royalty-derden per naam
        extra_namen_royalty: set = set()
        for kd in (rt, ws_d, b2b_k):
            for ed in (kd.get("extra_derden_per_naam") or []):
                if ed.get("type") == "royalty":
                    extra_namen_royalty.add(ed.get("naam", "Extra"))
        for nm in sorted(extra_namen_royalty):
            def _ev(kd, n=nm):
                return sum(x["bedrag"] for x in (kd.get("extra_derden_per_naam") or []) if x["type"] == "royalty" and x["naam"] == n)
            add_row(f"  {nm}",
                    -_ev(rt), -_ev(ws_d), -_ev(b2b_k),
                    -(_ev(rt)*verd_rt + _ev(ws_d)*verd_ws + _ev(b2b_k)*verd_b2b))

        add_if_nonzero("Overige kosten", "overige_kosten")
        add_row("Brutowinst",
                rt.get("brutowinst", 0), ws_d.get("brutowinst", 0), b2b_k.get("brutowinst", 0), gew("brutowinst"),
                bold=True, bg=C_GRIJS_LICHT)

        # Winstdeling-derden onder brutowinst
        if gew("auteur_winstdeling") > 0:
            add_if_nonzero("Auteur winstdeling", "auteur_winstdeling")
        if _is_wd("vertaler") and gew("vertaler") > 0:
            add_row("  Vertaler (winstdeling)",
                    -rt.get("vertaler", 0), -ws_d.get("vertaler", 0), -b2b_k.get("vertaler", 0), -gew("vertaler"))
        if _is_wd("illustrator") and gew("illustrator") > 0:
            add_row("  Illustrator (winstdeling)",
                    -rt.get("illustrator", 0), -ws_d.get("illustrator", 0), -b2b_k.get("illustrator", 0), -gew("illustrator"))
        if _is_wd("agent") and gew("agent") > 0:
            add_row("  Agent (winstdeling)",
                    -rt.get("agent", 0), -ws_d.get("agent", 0), -b2b_k.get("agent", 0), -gew("agent"))

        extra_namen_wd: set = set()
        for kd in (rt, ws_d, b2b_k):
            for ed in (kd.get("extra_derden_per_naam") or []):
                if ed.get("type") == "winstdeling":
                    extra_namen_wd.add(ed.get("naam", "Extra"))
        for nm in sorted(extra_namen_wd):
            def _evw(kd, n=nm):
                return sum(x["bedrag"] for x in (kd.get("extra_derden_per_naam") or []) if x["type"] == "winstdeling" and x["naam"] == n)
            add_row(f"  {nm} (winstdeling)",
                    -_evw(rt), -_evw(ws_d), -_evw(b2b_k),
                    -(_evw(rt)*verd_rt + _evw(ws_d)*verd_ws + _evw(b2b_k)*verd_b2b))

        add_row("Netto winst Maven",
                rt.get("netto_winst_maven", 0), ws_d.get("netto_winst_maven", 0), b2b_k.get("netto_winst_maven", 0), gew("netto_winst_maven"),
                bold=True, bg=C_GRIJS_LICHT)
        add_row("Marge %",
                rt.get("marge_pct", 0), ws_d.get("marge_pct", 0), b2b_k.get("marge_pct", 0), 0,
                bold=True, is_marge=True)

        if ti.get("heeft_partner") and gew("partner_winstdeling") > 0:
            add_row("  Partner-winstdeling (informatief)",
                    -rt.get("partner_winstdeling", 0), -ws_d.get("partner_winstdeling", 0),
                    -b2b_k.get("partner_winstdeling", 0), -gew("partner_winstdeling"))

        wf_tbl = Table(wf_rows, colWidths=[lbl_w, val_w, val_w, val_w, val_w])
        base_style = [
            ("FONTNAME", (0, 0), (-1, 0), FONT_B),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), C_GRIJS_LICHT),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_WIT, colors.HexColor("#FAFAFA")]),
        ] + row_styles
        wf_tbl.setStyle(TableStyle(base_style))
        story.append(wf_tbl)
        spacer(5)

    # ── 4. OPLAGE SIMULATIE ──
    # Gebruik de simulate/oplage API-logica rechtstreeks via run_calculation data
    try:
        from flask import current_app
        with current_app.test_client() as client:
            sim_resp = client.post(
                "/calculatie/api/simulate/oplage",
                json=data,
                content_type="application/json",
            )
            sim_data = sim_resp.get_json() if sim_resp.status_code == 200 else None
    except Exception:
        sim_data = None

    if sim_data and sim_data.get("rows"):
        section_header("OPLAGE SIMULATIE")

        sim_hdrs = ["Oplage", "Netto omzet", "Drukkosten", "Eenm. kosten", "Dealkosten", "Netto resultaat", "Marge %"]
        sim_rows = [sim_hdrs]
        sim_styles = []
        si = 1
        for row in sim_data["rows"]:
            is_be = row.get("is_break_even", False)
            bg = C_GROEN_LICHT if is_be else C_WIT
            vol = row.get("oplage", 0)
            lbl = f"{vol:,}".replace(",", ".") + ("  ← break-even" if is_be else "")
            marge_v = row.get("marge_pct", 0)
            sim_rows.append([
                Paragraph(lbl, sBold if is_be else sNorm),
                Paragraph(f"€ {row.get('omzet', 0):,.0f}".replace(",", "."), sRight),
                Paragraph(f"€ {row.get('kosten', 0) - row.get('eenmalig', 0) - row.get('dealkosten', 0):,.0f}".replace(",", "."), sRight),
                Paragraph(f"€ {row.get('eenmalig', 0):,.0f}".replace(",", "."), sRight),
                Paragraph(f"€ {row.get('dealkosten', 0):,.0f}".replace(",", "."), sRight),
                Paragraph(f"€ {row.get('netto_resultaat', 0):,.0f}".replace(",", "."), sBoldRight if is_be else sRight),
                Paragraph(
                    f'<font color="{marge_kleur(marge_v).hexval()}"><b>{marge_v*100:.1f}%</b></font>',
                    sRight
                ),
            ])
            if bg != C_WIT:
                sim_styles.append(("BACKGROUND", (0, si), (-1, si), bg))
            si += 1

        sim_col_w = usable / 7
        sim_lbl_w = usable - 6 * sim_col_w
        sim_tbl = Table(sim_rows, colWidths=[sim_lbl_w] + [sim_col_w] * 6)
        sim_tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), FONT_B),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), C_GRIJS_LICHT),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_WIT, colors.HexColor("#FAFAFA")]),
        ] + sim_styles))
        story.append(sim_tbl)
        spacer(3)
        note = Paragraph(
            "Dealkosten = voorschotten + royalty's en commissies aan alle partijen, inclusief winstdeling.",
            ps("note", fontSize=7, textColor=colors.HexColor("#888888")),
        )
        story.append(note)

    doc.build(story)
    buf.seek(0)

    titel_slug = (ti.get("titel", "calculatie") or "calculatie").replace(" ", "_")[:30]
    filename = f"calculatie_{titel_slug}.pdf"

    from flask import send_file
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )


# ── CSV Import ──

@bp.route("/api/import/csv", methods=["POST"])
def import_csv_file():
    """Import titels vanuit een CSV-bestand.

    Meerdere rijen met hetzelfde ISBN (of dezelfde titel+auteur) worden
    samengevoegd tot één titel met meerdere drukken.

    Verplicht: titel
    Aanbevolen: auteur, isbn, druknummer, oplage, drukkosten_per_ex,
                verkoopprijs_incl_btw, boekhandelskorting, auteur_winstdeling_pct
    """
    if "file" not in request.files:
        abort(400, description="Geen bestand meegegeven")

    file = request.files["file"]
    if not file.filename or not file.filename.endswith(".csv"):
        abort(400, description="Alleen .csv-bestanden toegestaan")

    raw = file.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(raw), delimiter=";")

    # Also try comma delimiter if semicolon yields no columns
    fieldnames = reader.fieldnames or []
    if len(fieldnames) <= 1:
        reader = csv.DictReader(io.StringIO(raw), delimiter=",")

    def parse_float(row, key, default=0.0, *aliases):
        for k in [key] + list(aliases):
            v = row.get(k, "")
            if v:
                try:
                    return float(v.replace(",", "."))
                except (ValueError, TypeError):
                    pass
        return default

    # Group rows by ISBN (or titel+auteur if no ISBN) to support multiple drukken
    from collections import OrderedDict
    groups = OrderedDict()  # key -> (first_row, [druk_rows])

    for row in reader:
        row = {k.strip().lower(): v.strip() for k, v in row.items() if k}
        titel = row.get("titel", "").strip()
        if not titel:
            continue

        isbn = row.get("isbn", "").strip()
        group_key = isbn if isbn else f"{titel}|{row.get('auteur', '')}"

        if group_key not in groups:
            groups[group_key] = (row, [])
        groups[group_key][1].append(row)

    count = 0
    for group_key, (first_row, druk_rows) in groups.items():
        titel = first_row.get("titel", "").strip()
        auteur = first_row.get("auteur", "")
        isbn = first_row.get("isbn", "")

        # Build drukken list from all rows in group, sorted by druknummer
        drukken = []
        for row in druk_rows:
            druk_raw = row.get("druknummer", row.get("druk", "1"))
            try:
                druknummer = max(1, int(druk_raw))
            except (ValueError, TypeError):
                druknummer = 1

            oplage = int(parse_float(row, "oplage", 2000, "oplage_1e_druk"))
            drukkosten = parse_float(row, "drukkosten_per_ex", 1.20, "drukkosten_1e_druk", "drukkosten")

            drukken.append({
                "druknummer": druknummer,
                "oplage": oplage,
                "drukkosten_per_ex": drukkosten,
                "kostenposten": [],
            })

        drukken.sort(key=lambda d: d["druknummer"])

        tid = storage.new_id()
        titel_data = {
            "titel_input": {
                "titel": titel,
                "auteur": auteur,
                "isbn": isbn,
                "verschijningsdatum": first_row.get("verschijningsdatum", ""),
                "verschenen": True,
                "verkoopprijs_incl_btw": parse_float(first_row, "verkoopprijs_incl_btw", 20.0),
                "btw_percentage": parse_float(first_row, "btw_percentage", 0.09),
                "boekhandelskorting": parse_float(first_row, "boekhandelskorting", 0.48),
                "drukken": drukken,
                "transactiekosten_pct": 0.002,
                "fulfillment_per_ex": 4.50,
                "cac_per_ex": 0.0,
                "distributie_cb_per_ex": 1.10,
                "b2b_porto_per_ex": 0.0,
                "b2b_korting_pct": 0.0,
                "auteur_winstdeling_pct": parse_float(first_row, "auteur_winstdeling_pct", 0.50),
                "auteur_royalty_staffel": [],
                "auteur_voorschot": 0,
                "agent_pct": 0.0, "agent_staffel": [], "agent_winstdeling_pct": 0.0, "agent_voorschot": 0,
                "vertaler_pct": 0.0, "vertaler_staffel": [], "vertaler_winstdeling_pct": 0.0, "vertaler_voorschot": 0,
                "illustrator_pct": 0.0, "illustrator_staffel": [], "illustrator_winstdeling_pct": 0.0, "illustrator_voorschot": 0,
                "heeft_partner": False,
                "partner_naam": "",
                "partner_winstdeling_pct": 0.5,
                "overige_kosten_pct": 0.0,
                "overige_kosten_items": [],
                "extra_derden": [],
            },
            "verdeling_webshop": 0.10,
            "verdeling_retail": 0.85,
            "verdeling_b2b": 0.05,
            "archived": False,
        }
        storage.save_titel(tid, titel_data)
        count += 1

    return jsonify({"imported": count})


# ── CSV Export ──

@bp.route("/api/export/csv", methods=["POST"])
def export_csv():
    data = request.get_json()
    res = run_calculation(data)

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")

    writer.writerow(["Maven Calculatie", res["titel"]])
    writer.writerow([])

    for druk in res["drukken"]:
        writer.writerow([druk["druk_type"].upper(), f"Oplage: {druk['oplage']}"])
        writer.writerow(["Kostenposten totaal", f"{druk.get('kosten_totaal', 0):.2f}"])
        writer.writerow(["Drukkosten totaal", f"{druk.get('drukkosten_totaal', 0):.2f}"])
        writer.writerow(["", "Webshop", "", "Retail (CB)", "", "B2B", ""])
        writer.writerow(["", "Bedrag", "Marge%", "Bedrag", "Marge%", "Bedrag", "Marge%"])

        writer.writerow([
            "Netto omzet",
            f"{druk['webshop']['netto_omzet']:.2f}", "",
            f"{druk['retail']['netto_omzet']:.2f}", "",
            f"{druk['b2b']['netto_omzet']:.2f}", "",
        ])

        writer.writerow([
            "Netto winst Maven",
            f"{druk['webshop']['netto_winst_maven']:.2f}",
            f"{druk['webshop']['marge_pct']:.1%}",
            f"{druk['retail']['netto_winst_maven']:.2f}",
            f"{druk['retail']['marge_pct']:.1%}",
            f"{druk['b2b']['netto_winst_maven']:.2f}",
            f"{druk['b2b']['marge_pct']:.1%}",
        ])

        writer.writerow([
            "Gewogen marge",
            f"{druk['gewogen_netto_winst']:.2f}",
            f"{druk['gewogen_marge_pct']:.1%}",
        ])
        writer.writerow([])

    output.seek(0)
    filename = f"calculatie_{res['titel'].replace(' ', '_')}.csv"

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
