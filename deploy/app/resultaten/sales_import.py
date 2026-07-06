"""
Sales-import (self-service) — verkoopdata per titel/kanaal → res_sales_snapshot.

De gedeployede app kan de sales-MCP niet zelf aanroepen, dus levert Hugo de
verkoopcijfers als export (CSV of .xlsx). Deze parser is bewust soepel: hij
herkent kolommen op alias en vult ``sales_sync.upsert_snapshot`` (idempotent op
isbn×kanaal×bron×jaar×kwartaal). Jaar/kwartaal mogen in het bestand staan óf als
formulier-veld meekomen (dan geldt dat voor het hele bestand — één kwartaal per
upload).

Herkende kolommen (case-insensitief, eerste match telt):
  isbn        : isbn, ean
  titel_naam  : titel, titel_naam, naam
  vorm        : vorm, verschijningsvorm, editie
  bron        : bron, kanaal, source
  jaar        : jaar, year
  kwartaal    : kwartaal, kwartaal_nr, q, periode  ('2026-Q2' of '2' of 'Q2')
  stuks       : stuks, verkocht, aantal, verkopen, saldo
  omzet       : omzet, netto, netto_omzet, netto_excl, bedrag
"""

import csv
import re

from . import sales_sync

# Geldige ISBN begint met 12+ cijfers; een suffix (bv. CB's "-A"-editievariant)
# mag blijven staan. Zo verliezen we geen verkoop op zulke edities, en kop-/
# totaalrijen (geen leidende cijfers) vallen er vanzelf uit.
ISBN_RE = re.compile(r"\d{12,}")


ALIASSEN = {
    "isbn": ("isbn", "ean"),
    "titel_naam": ("titel_naam", "titel", "naam"),
    "vorm": ("vorm", "verschijningsvorm", "editie"),
    "bron": ("bron", "kanaal", "source"),
    "jaar": ("jaar", "year"),
    "kwartaal": ("kwartaal", "kwartaal_nr", "periode", "q"),
    "stuks": ("stuks", "verkocht", "aantal", "verkopen", "saldo"),
    "omzet": ("omzet", "netto_omzet", "netto_excl", "netto", "bedrag"),
}


def _norm(s) -> str:
    return str(s or "").strip().lower().replace(" ", "_").replace(".", "")


def _kolom_map(headers: list[str]) -> dict:
    """{veld: kolomindex} op basis van de aliassen."""
    genormaliseerd = [_norm(h) for h in headers]
    uit = {}
    for veld, aliassen in ALIASSEN.items():
        for alias in aliassen:
            if alias in genormaliseerd:
                uit[veld] = genormaliseerd.index(alias)
                break
    return uit


def _parse_kwartaal(waarde, fallback: int | None) -> int | None:
    """'2026-Q2' → 2, 'Q3' → 3, '4' → 4; leeg → fallback."""
    s = str(waarde or "").strip().upper()
    if not s:
        return fallback
    if "-Q" in s:
        s = s.split("-Q", 1)[1]
    s = s.lstrip("Q")
    return int(s) if s.isdigit() else fallback


def _parse_jaar(waarde, fallback: int | None) -> int | None:
    s = str(waarde or "").strip()
    if "-Q" in s:
        s = s.split("-Q", 1)[0]
    return int(s) if s[:4].isdigit() and len(s) >= 4 else fallback


def _getal(waarde) -> float:
    s = str(waarde or "").strip().replace("€", "").replace(" ", "")
    if not s:
        return 0.0
    # NL-notatie: 1.234,56 → 1234.56
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _lees_rijen(path: str, filename: str) -> list[list]:
    """Lees CSV of .xlsx → lijst van rijen (lijsten)."""
    if filename.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        return [row for row in csv.reader(f, dialect)]


def import_sales(path: str, filename: str, *, jaar: int | None = None,
                 kwartaal: int | None = None) -> dict:
    """Importeer een sales-export. ``jaar``/``kwartaal`` gelden als het bestand
    die kolommen niet heeft (één kwartaal per upload)."""
    rijen = [r for r in _lees_rijen(path, filename) if r and any(c is not None and str(c).strip() for c in r)]
    if not rijen:
        raise ValueError("Leeg bestand.")

    # Headerrij = eerste rij die minstens isbn + (stuks of omzet) herkent.
    hdr_idx = kmap = None
    for i, r in enumerate(rijen[:10]):
        m = _kolom_map([str(c) for c in r])
        if "isbn" in m and ("stuks" in m or "omzet" in m):
            hdr_idx, kmap = i, m
            break
    if kmap is None:
        raise ValueError("Geen herkenbare kolommen (verwacht o.a. 'isbn' en 'stuks'/'omzet').")

    def cel(row, veld):
        i = kmap.get(veld)
        return row[i] if i is not None and i < len(row) else None

    out = []
    for r in rijen[hdr_idx + 1:]:
        isbn = str(cel(r, "isbn") or "").strip()
        if not ISBN_RE.match(isbn):                   # kop/totaal/lege rij overslaan
            continue
        jr = _parse_jaar(cel(r, "jaar"), jaar)
        kw = _parse_kwartaal(cel(r, "kwartaal"), kwartaal)
        if not jr or not kw:
            raise ValueError("Jaar/kwartaal ontbreekt — zet ze in het bestand of kies ze bij de upload.")
        out.append({
            "isbn": isbn,
            "titel_naam": str(cel(r, "titel_naam") or "").strip(),
            "vorm": str(cel(r, "vorm") or "").strip(),
            "bron": str(cel(r, "bron") or "").strip(),
            "jaar": jr, "kwartaal": kw,
            "stuks": int(_getal(cel(r, "stuks"))),
            "omzet": round(_getal(cel(r, "omzet")), 2),
        })

    if not out:
        raise ValueError("Geen geldige verkoopregels gevonden (ISBN's van 12+ cijfers).")
    res = sales_sync.upsert_snapshot(out)
    return res
