"""
SFP-historie import.

Leest een SFP-export (omzet/uitgaven per titel) en vult ``res_historie``: één
rij per ISBN met cumulatief verkocht (saldo) + cumulatieve netto-omzet t/m de
cutover-datum. SFP groepeert de edities al onder één titelnaam — die nemen we
over als groepssleutel.

Voorlopig leest deze de .xlsx-export (openpyxl, al aanwezig voor de
Excel-export). De CSV-variant die Hugo kan exporteren is een triviale aanpassing
(zelfde kolommen).

Kolommen in de SFP-export (0-indexed):
  0 ISBN · 1 titel · 2 verschijningsvorm · 7 saldo-verkopen · 11 netto excl.
Kop- en totaalregels worden overgeslagen door alleen rijen met een geldig
ISBN in kolom 0 te nemen.
"""

from datetime import datetime

from .models import Historie
from ..db import db


def parse_sfp_export(path):
    """Parse een SFP .xlsx-export → lijst dicts per editie/ISBN."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    for r in ws.iter_rows(values_only=True):
        isbn = str(r[0]).strip() if r and r[0] is not None else ""
        if isbn.isdigit() and len(isbn) >= 12:   # echte editie-regel, geen kop/totaal
            rows.append({
                "isbn": isbn,
                "titel_naam": (r[1] or "").strip(),
                "verschijningsvorm": (r[2] or "").strip(),
                "cumulatief_stuks": int(r[7] or 0),
                "cumulatief_netto_omzet": round(float(r[11] or 0), 2),
            })
    return rows


def import_sfp_historie(path, cutover_datum, import_batch=None):
    """Importeer een SFP-export in res_historie (idempotent per isbn+cutover)."""
    batch = import_batch or datetime.utcnow().strftime("sfp-%Y%m%d%H%M%S")
    rows = parse_sfp_export(path)

    n_new = n_upd = 0
    for row in rows:
        rec = Historie.query.filter_by(
            isbn=row["isbn"], cutover_datum=cutover_datum,
        ).first()
        if rec is None:
            rec = Historie(isbn=row["isbn"], cutover_datum=cutover_datum)
            db.session.add(rec)
            n_new += 1
        else:
            n_upd += 1
        rec.titel_naam = row["titel_naam"]
        rec.verschijningsvorm = row["verschijningsvorm"]
        rec.cumulatief_stuks = row["cumulatief_stuks"]
        rec.cumulatief_netto_omzet = row["cumulatief_netto_omzet"]
        rec.bron = "SFP"
        rec.import_batch = batch
        rec.imported_at = datetime.utcnow()

    db.session.commit()
    return {"batch": batch, "rijen": len(rows), "nieuw": n_new, "bijgewerkt": n_upd}
