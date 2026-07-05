"""
Exact-import (deterministische laag).

Leest een Exact FinTransactions-export en zet elke regel in ``res_kosten_geboekt``:
- koppelt aan een titel via de **Kostenplaats** (code = laatste 8 ISBN-cijfers),
- mapt het **grootboek** deterministisch naar een kostenstroom + grove categorie.

Dit is de basis; de fijnere match (welke calculatie-post precies, bv. "sticker")
en de twijfelgevallen gaan straks via de LLM-laag bovenop deze velden
(``calculatie_post`` / ``match_bron`` / ``match_confidence``).
"""

from datetime import datetime

from .models import KostenGeboekt
from ..db import db


# Grootboeknummer (eerste token) → (stroom, grove categorie).
GROOTBOEK_MAP = {
    "7030": ("kosten_per_ex", "productie"),     # All-in productiekosten
    "7005": ("kosten_per_ex", "productie"),     # Drukwerk
    "7003": ("kosten_per_ex", "productie"),     # Vormgeving
    "7016": ("kosten_per_ex", "productie"),     # Print on demand
    "7002": ("kosten_per_ex", "productie"),     # Redactie
    "7070": ("kosten_per_ex", "productie"),     # Inkopen overig
    "7007": ("kosten_per_ex", "campagne"),      # Promotie / Reclame
    "7115": ("kosten_per_ex", "campagne"),      # Verkoopkosten
    "7100": ("vast", "overig"),                 # Kosten internet/web
    "7031": ("vast", "cb-distributie"),         # Kostprijs CB distributie
    "7120": ("vast", "fulfillment"),            # Verzendkosten
    "4900": ("royalty", "auteur/agent"),        # Royalty's
}


def _periode(datum):
    """'2026-03-30 ...' → '2026-Q1'."""
    try:
        d = datum if isinstance(datum, datetime) else datetime.fromisoformat(str(datum)[:10])
        return f"{d.year}-Q{(d.month - 1) // 3 + 1}"
    except Exception:
        return ""


def _isbn_uit_kostenplaats(code):
    """Kostenplaats-code (laatste 8 ISBN-cijfers) → volledige Maven-ISBN.

    Maven-ISBN's beginnen met 97894…; we plakken die prefix ervoor. Leeg =
    geen titel (overhead → telt alleen in Maven-totaal).
    """
    code = str(code or "").strip()
    if code.isdigit() and len(code) == 8:
        return "97894" + code
    return ""


def parse_exact_export(path):
    """Parse een Exact .xlsx FinTransactions-export → lijst dicts per regel."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # Échte headerrij vinden: bevat álle kernkolommen (niet het criteria-blok
    # bovenaan, waar 'Grootboekrekening' ook losstaand voorkomt). Daarna lezen
    # we op vaste kolomposities — het Exact-exportformaat ligt vast.
    KERN = {"Bkst.nr.", "Bedrag", "Grootboekrekening", "Kostenplaats"}
    hdr_idx = next((i for i, r in enumerate(rows)
                    if r and KERN <= {str(c).strip() for c in r if c is not None}), None)
    if hdr_idx is None:
        raise ValueError("Geen FinTransactions-header gevonden — is dit een Exact-export?")

    NR, BKST, DATUM, RELATIE, OMSCHR, BEDRAG, GRB, KPLAATS = 0, 1, 2, 5, 6, 10, 14, 15

    out = []
    for r in rows[hdr_idx + 1:]:
        if not r or r[NR] is None:
            continue
        grb = str(r[GRB] or "").strip()
        if not grb:
            continue
        nummer = grb.split(" ")[0]            # "7030 - All-in…" → "7030"
        stroom, categorie = GROOTBOEK_MAP.get(nummer, ("overig", "overig"))
        try:
            bedrag = round(float(r[BEDRAG] or 0), 2)
        except (TypeError, ValueError):
            bedrag = 0.0
        out.append({
            "exact_ref": f"{r[BKST]}/{r[NR]}",
            "datum": str(r[DATUM] or "")[:10],
            "periode": _periode(r[DATUM]),
            "isbn": _isbn_uit_kostenplaats(r[KPLAATS]),
            "relatie": str(r[RELATIE] or "").strip(),
            "omschrijving": str(r[OMSCHR] or "").strip(),
            "grootboek": grb,
            "stroom": stroom,
            "categorie": categorie,
            "bedrag": bedrag,
        })
    return out


def import_exact(path, import_batch=None):
    """Importeer een Exact-export in res_kosten_geboekt (idempotent op exact_ref)."""
    batch = import_batch or datetime.utcnow().strftime("exact-%Y%m%d%H%M%S")
    rows = parse_exact_export(path)

    from .models import DispositieRegel
    regels = {r.relatie: r.dispositie for r in DispositieRegel.query.all()}

    n_new = n_upd = 0
    for row in rows:
        rec = KostenGeboekt.query.filter_by(exact_ref=row["exact_ref"]).first()
        nieuw = rec is None
        if nieuw:
            rec = KostenGeboekt(exact_ref=row["exact_ref"])
            db.session.add(rec)
            n_new += 1
        else:
            n_upd += 1
        # Handmatige titel-koppeling (herkoppel) niet overschrijven bij her-import:
        # de export heeft geen kostenplaats voor die regel, maar de mens wél.
        behoud_isbn = (not nieuw and rec.match_bron == "mens" and rec.isbn and not row.get("isbn"))
        for k, v in row.items():
            if k == "exact_ref":
                continue
            if k == "isbn" and behoud_isbn:
                continue
            setattr(rec, k, v)
        if not behoud_isbn:
            rec.match_bron = "regel"      # deterministisch via grootboek
        # Onthouden dispositie (per relatie) toepassen op regels zonder titel —
        # maar nooit over een handmatige per-regel-keuze heen.
        if not rec.isbn and not rec.dispositie:
            onthouden = regels.get((rec.relatie or "").strip().lower())
            if onthouden:
                rec.dispositie = onthouden
        rec.import_batch = batch
        rec.imported_at = datetime.utcnow()

    db.session.commit()
    return {"batch": batch, "rijen": len(rows), "nieuw": n_new, "bijgewerkt": n_upd}
